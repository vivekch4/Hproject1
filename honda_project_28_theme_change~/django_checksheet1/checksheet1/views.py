from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO
import json
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.timezone import localdate, localtime, now, make_aware
from django.views.decorators.csrf import csrf_exempt
from reportlab.pdfgen import canvas

from .models import (
    CheckSheet,
    CheckSheetImage,
    CustomUser,
    FilledCheckSheet,
    FilledStarterSheet,
    FormRequest,
    POCReadStatus,
    POCUpload,
    PageAccess,
    PasswordResetRequest,
    ProductionTarget,
    ProductionDb,
    Shifttime,
    StarterSheet,
    StarterZone,
    Zone,
    RejectionAlertConfig,
    RejectReason
)
from django.db.models import Q, F
import pytz
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from reportlab.lib.pagesizes import letter
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from PIL import Image as PILImage
from dateutil.parser import parse as parse_date, ParserError
from django.db import transaction
from django.views.decorators.http import require_http_methods



User = get_user_model()

connection_status = False

# import json
# import threading
# import ssl
# import datetime
# from django.http import HttpResponse
# import paho.mqtt.client as mqtt
# from .models import ProductionDb

# # MQTT broker settings
# BROKER = "broker.emqx.io"
# TOPIC = "myhome/test1"
# TCP_PORT = 1883
# SSL_PORT = 8883
# use_ssl = False  # Set to True if using TLS

# # Flag to avoid starting multiple clients
# mqtt_client_running = False


# def start_mqtt_client():
#     def on_connect(client, userdata, flags, rc):
#         if rc == 0:
#             print("Connected to MQTT Broker!")
#             client.subscribe(TOPIC)
#         else:
#             print(f"Failed to connect, return code {rc}")

#     def on_message(client, userdata, msg):
#         try:
#             data = json.loads(msg.payload.decode())
#             payload = data.get("payload", [])
#             print(payload, "payload")
#             if not payload:
#                 return

#             status = payload[0].get("status", {})
#             reg2 = status.get("Reg-2")
#             ts = status.get("timestamp")

#             if reg2 is not None and ts is not None:
#                 dt = dst.fromtimestamp(ts)
#                 if not ProductionDb.objects.filter(timestamp=dt).exists():
#                     ProductionDb.objects.create(
#                         Production_count=str(reg2), timestamp=dt
#                     )
#                     print(f"Saved: Reg-2={reg2}, Time={dt}")
#         except Exception as e:
#             print(f"Error processing message: {e}")

#     client = mqtt.Client()
#     client.on_connect = on_connect
#     client.on_message = on_message

#     if use_ssl:
#         client.tls_set(certfile=None, keyfile=None, tls_version=ssl.PROTOCOL_TLSv1_2)
#         client.tls_insecure_set(True)
#         port = SSL_PORT
#     else:
#         port = TCP_PORT

#     client.connect(BROKER, port, 60)
#     client.loop_forever()


# # Django view to start the MQTT listener
# def start_mqtt():
#     global mqtt_client_running
#     if not mqtt_client_running:
#         mqtt_client_running = True
#         threading.Thread(target=start_mqtt_client, daemon=True).start()


# start_mqtt()


# ----------------------------------------- login functions  --------------------------------#
def redirect_to_login(request):
    return redirect("login")


def is_shift_incharge(user):
    return user.is_authenticated and user.role == "shift_incharge"


def is_admin_or_incharge(user):
    return user.is_authenticated and (
        user.role in ["admin", "shift_incharge", "quality_incharge"]
    )


def is_quality_incharge(user):
    return user.is_authenticated and user.role == "quality_incharge"


def is_operator(user):
    return user.is_authenticated and user.role == "operator"


def is_admin(user):
    return user.is_authenticated and user.role == "admin"


def has_page_access(user, page_name):
    return PageAccess.objects.filter(
        user=user, page_name=page_name, has_access=True
    ).exists()

import random
import string
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.utils import timezone
from datetime import timedelta
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import OTP, CustomUser

import json

def generate_otp(length=6):
    """Generate a random 6-digit OTP."""
    return ''.join(random.choices(string.digits, k=length))


import random
import string
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.utils import timezone
from datetime import timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
from django.conf import settings
from .models import CustomUser, OTP  # Ensure these are imported

def login_view(request):
    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        password = request.POST.get("password")

        user = authenticate(request, employee_id=employee_id, password=password)

        if user and user.is_authenticated:
            if not user.is_active:
                return render(
                    request,
                    "checksheet/login.html",
                    {"error": "Your account is deactivated. Please contact an admin."}
                )

            # Store user info in session immediately
            request.session['otp_user_id'] = user.id
            request.session['employee_id'] = employee_id
            request.session.save()

            # Generate and send OTP synchronously
            try:
                # Generate OTP
                otp_code = ''.join(random.choices(string.digits, k=6))
                expires_at = timezone.now() + timedelta(minutes=10)

                # Clear any existing OTPs for this user
                OTP.objects.filter(user=user).delete()

                # Create new OTP
                OTP.objects.create(user=user, otp_code=otp_code, expires_at=expires_at)

                # Email content
                subject = 'Your OTP Code for Login'
                message = f"""
                Hello {user.username},

                Your One-Time Password (OTP) for login is: {otp_code}

                This OTP is valid for 10 minutes only.

                If you didn't request this OTP, please ignore this email.

                Best regards,
                Your Security Team
                """

                # Send email
                email = user.email
                msg = MIMEMultipart()
                msg["From"] = settings.SMTP_EMAIL
                msg["To"] = email
                msg["Subject"] = subject
                msg.attach(MIMEText(message, "plain"))

                # Send via SMTP
                with smtplib.SMTP("smtp.gmail.com", 587) as server:
                    server.starttls()
                    server.login(settings.SMTP_EMAIL, settings.SMTP_APP_PASSWORD)
                    server.send_message(msg)

                print(f"OTP sent successfully to {email} for user {employee_id}")

            except smtplib.SMTPException as e:
                print(f"SMTP error sending OTP to user {employee_id}: {str(e)}")
                return render(
                    request,
                    "checksheet/login.html",
                    {"error": "Failed to send OTP. Please try again."}
                )
            except Exception as e:
                print(f"Unexpected error sending OTP to user {employee_id}: {str(e)}")
                return render(
                    request,
                    "checksheet/login.html",
                    {"error": "An error occurred. Please try again."}
                )

            # Redirect immediately after sending OTP
            return redirect('verify_otp')
        else:
            return render(
                request,
                "checksheet/login.html",
                {"error": "Invalid credentials"}
            )

    return render(request, "checksheet/login.html")


def verify_otp(request):
    if request.method == "POST":
        otp_code = request.POST.get("otp_code")
        user_id = request.session.get('otp_user_id')
        
        if not user_id:
            return render(
                request,
                "checksheet/verify_otp.html",
                {"error": "Session expired. Please login again.", "redirect_to_login": True}
            )
        
        try:
            user = CustomUser.objects.get(id=user_id)
            
            # Get the most recent valid OTP
            otp = OTP.objects.filter(
                user=user, 
                otp_code=otp_code,
                expires_at__gt=timezone.now()
            ).first()
            
            if not otp:
                return render(
                    request,
                    "checksheet/verify_otp.html",
                    {"error": "Invalid or expired OTP. Please try again or request a new one."}
                )
            
            # Check if OTP is expired
            if otp.is_expired():
                return render(
                    request,
                    "checksheet/verify_otp.html",
                    {"error": "OTP has expired. Please request a new one."}
                )
            
            # Log the user in
            user.backend = 'checksheet1.authentication.EmployeeIDBackend'
            login(request, user)
            
            # Clean up
            OTP.objects.filter(user=user).delete()
            request.session.pop('otp_user_id', None)
            request.session.pop('employee_id', None)
            request.session.save()
            
            # Redirect based on role
            if user.role == "admin":
                return redirect("home")
            elif user.role == "quality_incharge":
                return redirect("home")
            elif user.role == "shift_incharge":
                return redirect("home")
            elif user.role == "operator":
                return redirect("operator_dashboard")
            else:
                return redirect("login")
                
        except CustomUser.DoesNotExist:
            return render(
                request,
                "checksheet/verify_otp.html",
                {"error": "User not found. Please login again.", "redirect_to_login": True}
            )
        except Exception as e:
            print(f"Error during OTP verification: {str(e)}")
            return render(
                request,
                "checksheet/verify_otp.html",
                {"error": "An error occurred. Please try again."}
            )
    
    # Check if user session exists when loading the page
    user_id = request.session.get('otp_user_id')
    if not user_id:
        return redirect('login')
    
    return render(request, "checksheet/verify_otp.html")

def logout_view(request):
    logout(request)
    return redirect("login")
# ----------------------------------------- bashboard function  --------------------------------#


def calculate_production_stats():
    global connection_status  # Consider replacing with a model-based solution
    stats = {
        "production_count": 0,
        "total_rejects": 0,
        "actual_production": 0,
        "efficiency": 0,
        "connection_status": "Connected" if connection_status else "Not Connected",
    }

    try:
        # Get the latest ProductionDb record by timestamp
        latest_production = ProductionDb.objects.latest("timestamp")
        print(
            f"Latest ProductionDb: id={latest_production.id}, Production_count={latest_production.Production_count}, timestamp={latest_production.timestamp}"
        )
        try:
            production_count_value = latest_production.Production_count
            stats["production_count"] = int(production_count_value)
            print(f"Set production_count to: {stats['production_count']}")
        except (ValueError, TypeError) as e:
            print(
                f"Error converting Production_count '{production_count_value}' to int: {e}"
            )
            stats["production_count"] = 0
            stats["error"] = "Invalid production count format"
            return stats
    except ProductionDb.DoesNotExist:
        print("No production data found")
        stats["error"] = "No production data found"
        return stats

    try:
        # Calculate reject count
        reject_count = (
            FilledCheckSheet.objects.aggregate(
                total_rejects=Count("id", filter=Q(rejected_by_id__isnull=False))
            )["total_rejects"]
            or 0
        )
        print(f"Rejects from rejected_by_id: {reject_count}")

        # Handle status_data["completely_reject"]
        filled_entries = FilledCheckSheet.objects.filter(
            status_data__has_key="completely_reject"
        )
        extra_rejects = 0
        for entry in filled_entries:
            value = entry.status_data.get("completely_reject")
            if isinstance(value, int):
                extra_rejects += value
            else:
                try:
                    extra_rejects += int(value)
                except (ValueError, TypeError):
                    if value == "Yes" or (isinstance(value, str) and value.strip()):
                        extra_rejects += 1
        stats["total_rejects"] = reject_count + extra_rejects
        print(
            f"Extra rejects from status_data: {extra_rejects}, Total rejects: {stats['total_rejects']}"
        )
    except Exception as e:
        print(f"Error calculating rejects: {e}")
        stats["error"] = "Error calculating rejects"
        return stats

    # Ensure actual_production is non-negative
    stats["actual_production"] = max(
        0, stats["production_count"] - stats["total_rejects"]
    )
    if stats["production_count"] > 0:
        stats["efficiency"] = round(
            (stats["actual_production"] / stats["production_count"]) * 100, 2
        )
    else:
        stats["efficiency"] = 0

    print("Final calculated stats:", stats)
    return stats


def broadcast_production_update():
    print("Entering broadcast_production_update")
    try:
        data = calculate_production_stats()
        print(f"Prepared data for broadcast: {data}")

        channel_layer = get_channel_layer()
        if channel_layer is None:
            print("Error: Channel layer is None")
            return
        print("Channel layer retrieved successfully")

        async_to_sync(channel_layer.group_send)(
            "production_group", {"type": "send_production_update", "data": data}
        )
        print("Group send called successfully")
    except Exception as e:
        print(f"Error in broadcast_production_update: {str(e)}")


@login_required
def home(request):
    user = request.user
    today = timezone.now()  # Use timezone.now() instead of now()
    today_start = today.replace(hour=0, minute=0, second=0, microsecond=0)
    start_of_week = today - timezone.timedelta(days=today.weekday())

    # Get all unique lines from checksheets
    all_lines = CheckSheet.objects.values_list("line", flat=True).distinct()
    lines_list = list(all_lines)

    # Fetch the latest production target
    latest_target = ProductionTarget.objects.order_by('-updated_at').first()
    target_value = latest_target.target_value if latest_target else 0  # Default to 0 if no target exists

    if user.role == "admin" or has_page_access(user, "home"):
        checksheets = CheckSheet.objects.all()
        starter = StarterSheet.objects.all()

        today_yes_counts_json = json.dumps(
            get_yes_counts(checksheets, today_start, today)
        )
        weekly_yes_counts_json = json.dumps(
            get_yes_counts(checksheets, start_of_week, today)
        )

        # Create checksheet options with line data attribute
        checksheet_options = "".join(
            f'<option value="{sheet.id}" data-line="{sheet.line}">{sheet.name}</option>'
            for sheet in checksheets
        )

        checksheets_data_json = json.dumps(
            [
                {
                    "id": sheet.id,
                    "name": sheet.name,
                    "line": sheet.line,
                    "fields": get_checksheet_fields(sheet.id),
                }
                for sheet in checksheets
            ]
        )
    else:
        checksheets = CheckSheet.objects.filter(assigned_users=user)
        starter = StarterSheet.objects.filter(assigned_users=user)
        today_yes_counts_json = weekly_yes_counts_json = checksheets_data_json = "[]"
        checksheet_options = ""

    if user.role == "admin":
        # Admins see all pending starter sheets
        pending_acknowledgments = FilledStarterSheet.objects.filter(
            Q(approval_status="level_2_approved", requires_level_3_approval=True)
        ).exclude(approval_status="rejected")

        # Count starter sheets approved by admin (excluding rejected)
        approved_count = (
            FilledStarterSheet.objects.filter(level_3_approval_id=user.id)
            .exclude(approval_status="rejected")
            .count()
        )
    else:
        # Regular users see starter sheets they are assigned to approve
        pending_acknowledgments = FilledStarterSheet.objects.filter(
            Q(assigned_level_1_approver=user, approval_status="pending")
            | Q(assigned_level_2_approver=user, approval_status="level_1_approved")
        ).exclude(approval_status="rejected")

        # Count starter sheets approved by user (excluding rejected)
        approved_count = (
            FilledStarterSheet.objects.filter(
                Q(
                    level_1_approval_id=user.id,
                    approval_status__in=[
                        "level_1_approved",
                        "level_2_approved",
                        "completed",
                    ],
                )
                | Q(
                    level_2_approval_id=user.id,
                    approval_status__in=["level_2_approved", "completed"],
                )
            )
            .exclude(approval_status="rejected")
            .count()
        )

    # Get pending starter sheets count
    pending_count = pending_acknowledgments.count()

    # Convert to values for template
    pending_acknowledgments = pending_acknowledgments.values(
        "id",
        "startersheet__id",
        "startersheet__name",
        "filled_by__id",
        "filled_by__username",
        "shift",
        "timestamp",
        "approval_status",
        "rejection_reason",
        "rejected_by_id",
        "level_1_approval_id",
        "level_2_approval_id",
    ).order_by("timestamp")

    # Handle FilledCheckSheet data
    if user.role == "admin":
        # Admins see all pending check sheets requiring acknowledgment
        pending_check_acknowledgments = FilledCheckSheet.objects.filter(
            Q(
                send_acknowledgment=True,
                approval_status="level_2_approved",
                requires_level_3_approval=True,
            )
        )

        # Check sheets approved by admin (excluding rejected)
        approved_check_sheets = FilledCheckSheet.objects.filter(
            Q(send_acknowledgment=True, level_3_approval_id=user.id)
        ).exclude(approval_status="rejected")
    else:
        # Regular users see check sheets they are assigned to approve
        pending_check_acknowledgments = FilledCheckSheet.objects.filter(
            Q(send_acknowledgment=True)
            & (
                Q(assigned_level_1_approver=user, approval_status="pending")
                | Q(assigned_level_2_approver=user, approval_status="level_1_approved")
            )
        ).exclude(approval_status="rejected")

        # Check sheets approved by user (excluding rejected)
        approved_check_sheets = FilledCheckSheet.objects.filter(
            Q(send_acknowledgment=True)
            & (
                Q(
                    level_1_approval_id=user.id,
                    approval_status__in=[
                        "level_1_approved",
                        "level_2_approved",
                        "completed",
                    ],
                )
                | Q(
                    level_2_approval_id=user.id,
                    approval_status__in=["level_2_approved", "completed"],
                )
            )
        ).exclude(approval_status="rejected")

    # Count unique combinations for pending check sheets
    pending_unique_combinations = {
        (check.timestamp.date(), check.shift, check.user_id, check.checksheet_id)
        for check in pending_check_acknowledgments
    }

    # Count unique combinations for approved check sheets
    approved_unique_combinations = {
        (check.timestamp.date(), check.shift, check.user_id, check.checksheet_id)
        for check in approved_check_sheets
    }

    # Get consolidated counts
    check_pending_count = len(pending_unique_combinations)
    check_approved_count = len(approved_unique_combinations)

    return render(
        request,
        "checksheet/home.html",
        {
            "checksheets": checksheets,
            "Starter": starter,
            "today_yes_counts_json": today_yes_counts_json,
            "weekly_yes_counts_json": weekly_yes_counts_json,
            "checksheet_options": checksheet_options,
            "checksheets_data_json": checksheets_data_json,
            "pending_count": check_pending_count,
            "acknowledged_count": check_approved_count,
            "startersheet_pending_count": pending_count,
            "startersheet_acknowledged_count": approved_count,
            "lines_list": lines_list,
            "today": today,
            "target_value": target_value,  # Add target value to context
        },
    )
@login_required
def get_pie_chart_data(request):
    checksheet_id = request.GET.get("checksheet_id")
    time_period = request.GET.get("period", "week")
    selected_date = request.GET.get("date", None)
    print(f"Selected date: {selected_date}")

    if not checksheet_id:
        return JsonResponse({"error": "Checksheet ID is required"}, status=400)

    today = timezone.now()

    # Calculate start and end dates
    if selected_date:
        try:
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d")
            selected_date = timezone.make_aware(selected_date, timezone.get_current_timezone())
            start_date = selected_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            print(f"Query range: {start_date} to {end_date}")
        except ValueError:
            print("Invalid date format")
            return JsonResponse({"error": "Invalid date format"}, status=400)
    else:
        if time_period == "today":
            start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        elif time_period == "week":
            start_date = today - timezone.timedelta(days=today.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timezone.timedelta(days=6, hours=23, minutes=59, seconds=59)
        elif time_period == "month":
            start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = (start_date + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(microseconds=1)
        elif time_period == "year":
            start_date = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        else:
            start_date = today - timezone.timedelta(days=today.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timezone.timedelta(days=6, hours=23, minutes=59, seconds=59)
        print(f"Query range (period={time_period}): {start_date} to {end_date}")

    try:
        checksheet = CheckSheet.objects.get(id=checksheet_id)
        if (
            request.user.role != "admin"
            and request.user.role != "shift_incharge"
            and request.user.role != "quality_incharge"
            and not checksheet.assigned_users.filter(id=request.user.id).exists()
        ):
            return JsonResponse({"error": "Access denied"}, status=403)

        filled_sheets = FilledCheckSheet.objects.filter(
            checksheet=checksheet, timestamp__gte=start_date, timestamp__lte=end_date
        )
        print(f"Queryset: {filled_sheets}")
        print(f"Queryset count: {filled_sheets.count()}")
        for sheet in filled_sheets:
            print(f"Sheet timestamp: {sheet.timestamp}, Status: {sheet.status_data}")

        field_yes_counts = {}
        for filled in filled_sheets:
            status_data = filled.status_data
            for field, value in status_data.items():
                if field != "completely_reject":
                    if value == "Yes":
                        field_yes_counts[field] = field_yes_counts.get(field, 0) + 1
                    elif isinstance(value, int):
                        field_yes_counts[field] = field_yes_counts.get(field, 0) + value

        all_fields = get_checksheet_fields(checksheet_id)
        for field in all_fields:
            if field not in field_yes_counts:
                field_yes_counts[field] = 0

        print(f"{checksheet.name} {field_yes_counts} {time_period} pie chart")
        return JsonResponse(
            {
                "checksheet_name": checksheet.name,
                "zone_yes_counts": field_yes_counts,
                "time_period": time_period,
            }
        )
    except CheckSheet.DoesNotExist:
        return JsonResponse({"error": "Checksheet not found"}, status=404)


def get_checksheet_fields(checksheet_id):
    """Helper function to get field names for a checksheet"""
    try:
        # Get all filled checksheets for this checksheet
        filled_sheets = FilledCheckSheet.objects.filter(checksheet_id=checksheet_id)

        # Collect all unique field names
        fields = set()
        for sheet in filled_sheets:
            for field in sheet.status_data.keys():
                if field != "completely_reject":
                    fields.add(field)

        return sorted(list(fields))
    except Exception:
        return []

@login_required
def get_chart_data(request):
    period = request.GET.get("period", "today")
    line = request.GET.get("line", "all")
    selected_date = request.GET.get("date", None)
    today = timezone.now()

    # Determine date range
    if selected_date:
        try:
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d")
            selected_date = timezone.make_aware(selected_date, timezone.get_current_timezone())
            start_date = selected_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date.replace(hour=23, minute=59, second=59, microsecond=999999)
            print(f"Query range: {start_date} to {end_date}")
        except ValueError:
            print("Invalid date format")
            return JsonResponse({"error": "Invalid date format"}, status=400)
    else:
        if period == "today":
            start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timezone.timedelta(days=1)
        elif period == "week":
            start_date = today - timezone.timedelta(days=today.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timezone.timedelta(days=6, hours=23, minutes=59, seconds=59)
        elif period == "month":
            start_date = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = (start_date + timezone.timedelta(days=32)).replace(day=1) - timezone.timedelta(microseconds=1)
        elif period == "year":
            start_date = today.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
        else:
            return JsonResponse({"error": "Invalid period"}, status=400)
        print(f"Query range (period={period}): {start_date} to {end_date}")

    # Filter checksheets by line
    if line == "all":
        checksheets = CheckSheet.objects.all()
    else:
        checksheets = CheckSheet.objects.filter(line=line)

    yes_counts = get_yes_counts(checksheets, start_date, end_date)
    print(f"Yes counts: {yes_counts}")
    return JsonResponse(yes_counts, safe=False)

def get_yes_counts(checksheets, start_date, end_date):
    """Helper function to get yes counts for a given date range"""
    yes_counts = []

    print(f"Timestamps: {start_date} to {end_date}")
    for sheet in checksheets:
        filled_sheets = FilledCheckSheet.objects.filter(
            checksheet=sheet, timestamp__gte=start_date, timestamp__lte=end_date
        )

        print(f"Sheet: {sheet.name}, Count: {filled_sheets.count()}")
        for fs in filled_sheets[:3]:
            print(f"  - Timestamp: {fs.timestamp}")
            print(f"  - Status data: {fs.status_data}")

        yes_count = 0
        for filled in filled_sheets:
            status_data = filled.status_data
            for key, value in status_data.items():
                if key == "completely_reject":
                    continue
                if value == "Yes":
                    yes_count += 1
                    print(f"  - Found 'Yes' for key: {key}")
                elif isinstance(value, int):
                    yes_count += value
                    print(f"  - Found integer {value} for key: {key}")

        yes_counts.append({"name": sheet.name, "yes_count": yes_count})

    return yes_counts


@login_required
def get_dashboard_stats(request):
    line = request.GET.get("line", "all")
    selected_date = request.GET.get("date", None)
    today = timezone.now()

    # Determine date range
    if selected_date:
        try:
            selected_date = timezone.datetime.strptime(selected_date, "%Y-%m-%d")
            selected_date = timezone.make_aware(selected_date, timezone.get_current_timezone())
            start_date = selected_date.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = start_date + timezone.timedelta(days=1, microseconds=-1)
            print(f"Query range: {start_date} to {end_date}")
        except ValueError:
            print("Invalid date format")
            return JsonResponse({"error": "Invalid date format"}, status=400)
    else:
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_date + timezone.timedelta(days=1, microseconds=-1)
        print(f"Query range (today): {start_date} to {end_date}")

    # Filter checksheets by line
    if line == "all":
        checksheets = CheckSheet.objects.all()
    else:
        checksheets = CheckSheet.objects.filter(line=line)

    # Calculate stats
    filled_sheets = FilledCheckSheet.objects.filter(
        checksheet__in=checksheets, timestamp__gte=start_date, timestamp__lte=end_date
    )
    print(f"Queryset: {filled_sheets}")
    print(f"Queryset count: {filled_sheets.count()}")
    for sheet in filled_sheets:
        print(f"Sheet timestamp: {sheet.timestamp}, Status: {sheet.status_data}")

    total_rejects = 0

    # Count rejects from filled sheets
    for filled in filled_sheets:
        status_data = filled.status_data
        for key, value in status_data.items():
            if key == "completely_reject" and value == "Yes":
                total_rejects += 1

    # Get total production from last ProductionDb entry for the date
    try:
        last_production = ProductionDb.objects.filter(
            timestamp__gte=start_date,
            timestamp__lte=end_date
        ).latest('timestamp')
        total_production = int(last_production.Production_count)
    except ProductionDb.DoesNotExist:
        total_production = 0

    # Calculate actual production
    actual_production = total_production - total_rejects

    # Calculate efficiency
    efficiency = (
        (actual_production / total_production * 100) if total_production > 0 else 0
    )

    response_data = {
        "production_count": total_production,
        "total_rejects": total_rejects,
        "actual_production": actual_production,
        "efficiency": f"{efficiency:.2f}%",
    }
    print(f"Dashboard stats: {response_data}")
    return JsonResponse(response_data)



# ----------------------------------------- Create Checksheet--------------------------------#
@login_required
def create_checksheet(request):
    if request.user.role == "admin" or has_page_access(request.user, "all_checksheets"):
        checksheets = CheckSheet.objects.all()
        Starter = StarterSheet.objects.all()
        all_users = get_user_model().objects.all()

        if request.method == "POST":
            line = request.POST.get("line")
            checksheet_name = request.POST.get("checksheet_name")
            image_1 = request.FILES.get("image_1")
            image_2 = request.FILES.get("image_2")
            image_3 = request.FILES.get("image_3")
            image_4 = request.FILES.get("image_4")

            if checksheet_name:
                # Create the checksheet
                checksheet = CheckSheet.objects.create(
                    name=checksheet_name, line=line, created_by=request.user
                )

                # Save the available images (no need for all four)
                if image_1:
                    CheckSheetImage.objects.create(checksheet=checksheet, image=image_1)
                if image_2:
                    CheckSheetImage.objects.create(checksheet=checksheet, image=image_2)
                if image_3:
                    CheckSheetImage.objects.create(checksheet=checksheet, image=image_3)
                if image_4:
                    CheckSheetImage.objects.create(checksheet=checksheet, image=image_4)

                # Save Zones dynamically with input types
                zone_count = int(request.POST.get("zone_count", 0))
                for i in range(zone_count):
                    zone_name = request.POST.get(f"zone_{i}")
                    zone_type = request.POST.get(f"zone_type_{i}")

                    if zone_name and zone_type:
                        Zone.objects.create(
                            checksheet=checksheet,
                            name=zone_name,
                            input_type=zone_type,
                        )

                success = True
                messages.success(
                    request,
                    "Checksheet created successfully!",
                    extra_tags="checksheet_creation",
                )
            else:
                success = False

            return redirect("all_checksheets")

        return render(
            request,
            "checksheet/create_checksheet.html",
            {"checksheets": checksheets, "Starter": Starter, "all_users": all_users},
        )

    return render(request, "checksheet/access_denied.html")


# -----------------------------------------  All Checksheet --------------------------------#
@login_required
def all_checksheets(request):
    if request.user.role == "admin" or has_page_access(request.user, "all_checksheets"):
        if request.user.role == "admin" or has_page_access(
            request.user, "all_checksheets"
        ):
            # Admin sees all CheckSheets
            checksheets = CheckSheet.objects.prefetch_related(
                "zones", "assigned_users"
            ).all()
            Starter = StarterSheet.objects.all()
        else:
            # Operators see only assigned CheckSheets
            Starter = StarterSheet.objects.filter(assigned_users=request.user)
            checksheets = CheckSheet.objects.prefetch_related(
                "zones", "assigned_users"
            ).filter(assigned_users=request.user)
        all_users = get_user_model().objects.all()

        return render(
            request,
            "checksheet/all_checksheets.html",
            {"checksheets": checksheets, "Starter": Starter, "all_users": all_users},
        )

    return render(request, "checksheet/access_denied.html")


def get_unique_copy_name(original_name):
    base_name = f"{original_name} copy"
    counter = 1
    new_name = base_name
    while CheckSheet.objects.filter(name=new_name).exists():
        new_name = f"{base_name} {counter}"
        counter += 1
    return new_name

@login_required
def copy_checksheet(request, checksheet_id):
    if request.user.role != "admin" and not has_page_access(request.user, "all_checksheets"):
        return render(request, "checksheet/access_denied.html")

    original_checksheet = get_object_or_404(CheckSheet, id=checksheet_id)

    # Get a unique name for the copy
    new_name = get_unique_copy_name(original_checksheet.name)

    new_checksheet = CheckSheet.objects.create(
        name=new_name,
        line=original_checksheet.line,
        created_by=request.user,
        level_1_approver=original_checksheet.level_1_approver,
        level_2_approver=original_checksheet.level_2_approver,
        require_level_3_approval=original_checksheet.require_level_3_approval,
    )

    new_checksheet.assigned_users.set(original_checksheet.assigned_users.all())

    for zone in original_checksheet.zones.all():
        Zone.objects.create(
            checksheet=new_checksheet,
            name=zone.name,
            input_type=zone.input_type,
        )

    for image in original_checksheet.images.all():
        CheckSheetImage.objects.create(
            checksheet=new_checksheet,
            image=image.image,
        )

    messages.success(request, f'CheckSheet "{original_checksheet.name}" copied successfully as "{new_name}".', extra_tags='checksheet_creation')
    return redirect('all_checksheets')
import csv


def get_unique_starter_copy_name(original_name):
    base_name = f"{original_name} copy"
    counter = 1
    new_name = base_name
    while StarterSheet.objects.filter(name=new_name).exists():
        new_name = f"{base_name} {counter}"
        counter += 1
    return new_name
@login_required
def copy_startersheet(request, startersheet_id):
    if request.user.role != "admin" and not has_page_access(request.user, "all_startersheet"):
        return render(request, "startersheet/access_denied.html")

    original_startersheet = get_object_or_404(StarterSheet, id=startersheet_id)

    # Get a unique name for the copy
    new_name = get_unique_starter_copy_name(original_startersheet.name)

    new_startersheet = StarterSheet.objects.create(
        name=new_name,
        line=original_startersheet.line,
        created_by=request.user,
        level_1_approver=original_startersheet.level_1_approver,
        level_2_approver=original_startersheet.level_2_approver,
        require_level_3_approval=original_startersheet.require_level_3_approval,
    )

    new_startersheet.assigned_users.set(original_startersheet.assigned_users.all())

    for zone in original_startersheet.zones.all():
        StarterZone.objects.create(
            startersheet=new_startersheet,
            name=zone.name,
            type=zone.type,
            min_value=zone.min_value,
            max_value=zone.max_value,
            unit=zone.unit,
            check_method=zone.check_method,
            image=zone.image,
            standard=zone.standard,
        )

    # Copy associated POCUpload PDFs
    for poc in original_startersheet.assigned_pocs.all():
        new_poc = POCUpload.objects.create(
            pdf=poc.pdf,
        )
        new_poc.assigned_startersheets.add(new_startersheet)

    messages.success(request, f'StarterSheet "{original_startersheet.name}" copied successfully as "{new_name}".', extra_tags='StarterSheet_creation')
    return redirect('all_startersheet')

@login_required
def export_checksheets(request):
    if request.method == 'POST':
        import json
        checksheet_ids = json.loads(request.POST.get('checksheet_ids', '[]'))
        
        # Fetch selected checksheets
        checksheets = CheckSheet.objects.filter(id__in=checksheet_ids).prefetch_related(
            'zones', 'assigned_users', 'level_1_approver', 'level_2_approver'
        )

        # Create the CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="checksheets_export.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Name', 'Zones', 'Line', 'Assigned Users', 
            'Level 1 Approver', 'Level 2 Approver', 'Level 3 Approval Required'
        ])

        for checksheet in checksheets:
            zones = ', '.join([zone.name for zone in checksheet.zones.all()])
            users = ', '.join([user.username for user in checksheet.assigned_users.all()])
            level_1 = checksheet.level_1_approver.username if checksheet.level_1_approver else 'Not assigned'
            level_2 = checksheet.level_2_approver.username if checksheet.level_2_approver else 'Not assigned'
            level_3 = 'Yes' if checksheet.require_level_3_approval else 'No'

            writer.writerow([
                checksheet.id,
                checksheet.name,
                zones,
                checksheet.line,
                users,
                level_1,
                level_2,
                level_3
            ])

        return response

    return HttpResponse(status=400)

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import csv
import json
  # Adjust the import based on your models' location

@login_required
def export_startersheets(request):
    if request.method == 'POST':
        startersheet_ids = json.loads(request.POST.get('startersheet_ids', '[]'))
        
        # Fetch selected startersheets
        startersheets = StarterSheet.objects.filter(id__in=startersheet_ids).prefetch_related(
            'zones', 'assigned_users', 'assigned_pocs', 'level_1_approver', 'level_2_approver'
        )

        # Create the CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="startersheets_export.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Name', 'Parameters', 'Line', 'Assigned Users', 
            'Assigned OPS', 'Level 1 Approver', 'Level 2 Approver', 'Level 3 Approval Required'
        ])

        for startersheet in startersheets:
            zones = ', '.join([zone.name for zone in startersheet.zones.all()])
            users = ', '.join([user.username for user in startersheet.assigned_users.all()])
            pocs = ', '.join([poc.pdf.name[9:] for poc in startersheet.assigned_pocs.all()])  # Remove 'media/poc/' prefix
            level_1 = startersheet.level_1_approver.username if startersheet.level_1_approver else 'Not assigned'
            level_2 = startersheet.level_2_approver.username if startersheet.level_2_approver else 'Not assigned'
            level_3 = 'Yes' if startersheet.require_level_3_approval else 'No'

            writer.writerow([
                startersheet.id,
                startersheet.name,
                zones,
                startersheet.line,
                users,
                pocs,
                level_1,
                level_2,
                level_3
            ])

        return response

    return HttpResponse(status=400)
# ----------------------------------------- Update CheckSheet--------------------------------#


@login_required
def update_checksheet(request, checksheet_id):
    if not (request.user.role == "admin" or has_page_access(request.user, "all_checksheets")):
        return render(request, "checksheet/access_denied.html")

    checksheet = get_object_or_404(CheckSheet, id=checksheet_id)
    zones = checksheet.zones.all()
    images = list(checksheet.images.all()[:4])
    all_users = CustomUser.objects.all()  # Fetch all users for dropdowns
    while len(images) < 4:
        images.append(None)  # Fill empty slots with None

    if request.method == "POST":
        # Handle zone deletion
        delete_zone_id = request.POST.get("delete_zone")
        if delete_zone_id:
            try:
                zone = zones.get(id=delete_zone_id)
                zone.delete()
                messages.success(request, "Zone deleted successfully!", extra_tags="zone_delete")
                return redirect("update_checksheet", checksheet_id=checksheet_id)
            except Zone.DoesNotExist:
                messages.error(request, "Zone not found.", extra_tags="zone_delete_error")
                return redirect("update_checksheet", checksheet_id=checksheet_id)

        # Update checksheet name and line
        checksheet_name = request.POST.get("checksheet_name")
        line = request.POST.get("line")
        if checksheet_name:
            checksheet.name = checksheet_name
        if line:
            checksheet.line = line

        # Update assigned users
        user_ids = request.POST.getlist("user_ids")
        users = CustomUser.objects.filter(id__in=user_ids)
        checksheet.assigned_users.set(users)

        # Update Level 1 approver
        level_1_approver_id = request.POST.get("level_1_approver")
        checksheet.level_1_approver = CustomUser.objects.filter(id=level_1_approver_id).first() if level_1_approver_id else None

        # Update Level 2 approver
        level_2_approver_id = request.POST.get("level_2_approver")
        checksheet.level_2_approver = CustomUser.objects.filter(id=level_2_approver_id).first() if level_2_approver_id else None

        # Update Level 3 approval
        require_level_3 = request.POST.get("require_level_3") == "True"
        checksheet.require_level_3_approval = require_level_3

        checksheet.save()

        # Update existing images or save new ones
        for i in range(4):
            new_image = request.FILES.get(f"checksheet_image_{i+1}")
            if new_image:
                if images[i] is None:
                    CheckSheetImage.objects.create(checksheet=checksheet, image=new_image)
                else:
                    images[i].image = new_image
                    images[i].save()

        # Update existing zones
        for zone in zones:
            zone_name = request.POST.get(f"zone_{zone.id}")
            zone_type = request.POST.get(f"zone_type_{zone.id}")
            if zone_name and zone_type:
                zone.name = zone_name
                zone.input_type = zone_type
                zone.save()
            else:
                messages.warning(
                    request,
                    f"Zone {zone.name} was not updated due to missing name or type.",
                    extra_tags="zone_update_warning",
                )

        # Create new zones
        new_zone_names = request.POST.getlist("new_zone_names[]")
        new_zone_types = request.POST.getlist("new_zone_types[]")
        for name, input_type in zip(new_zone_names, new_zone_types):
            if name and input_type:
                Zone.objects.create(
                    checksheet=checksheet,
                    name=name,
                    input_type=input_type
                )
            else:
                messages.warning(
                    request,
                    "A new zone was not created due to missing name or type.",
                    extra_tags="zone_create_warning",
                )

        messages.success(request, "Checksheet updated successfully!", extra_tags="checksheet_update")
        return redirect("all_checksheets")

    return render(
        request,
        "checksheet/update_checksheet.html",
        {
            "checksheet": checksheet,
            "zones": zones,
            "images": images,
            "all_users": all_users,
        },
    )

    


# ----------------------------------------- Add Zone in CheckSheet--------------------------------#
# @login_required
# def add_zone(request, checksheet_id):
#     if request.user.role == "admin" or has_page_access(request.user, "all_checksheets"):
#         if request.user.role == "admin" or has_page_access(
#             request.user, "all_checksheets"
#         ):
#             # Admin sees all CheckSheets and StarterSheets
#             checksheets = CheckSheet.objects.all()
#             Starter = StarterSheet.objects.all()
#         else:
#             # Operators see only assigned CheckSheets and StarterSheets
#             checksheets = CheckSheet.objects.filter(assigned_users=request.user)
#             Starter = StarterSheet.objects.filter(assigned_users=request.user)
#         checksheet = get_object_or_404(CheckSheet, id=checksheet_id)

#         if request.method == "POST":
#             zone_name = request.POST.get("zone_name")
#             zone_type = request.POST.get("zone_type")  # Get selected type

#             if zone_name and zone_type:
#                 Zone.objects.create(
#                     checksheet=checksheet, name=zone_name, input_type=zone_type
#                 )
#                 messages.success(
#                     request,
#                     "Zone Added successfully!",
#                     extra_tags="zone_add",
#                 )
#                 # Ensure your Zone model has an `input_type` field

#             return redirect("all_checksheets")

#         return render(
#             request,
#             "checksheet/add_zone.html",
#             {"checksheet": checksheet, "Starter": Starter, "checksheets": checksheets},
#         )

#     return render(request, "checksheet/access_denied.html")


# ----------------------------------------- Fill Checksheet--------------------------------#


@login_required
def fill_checksheet(request, checksheet_id=None):
    if request.user.role in ["operator","admin"] or has_page_access(
        request.user, "fill_checksheet_detail"
    ):
        if request.user.role == "admin":
            checksheets = CheckSheet.objects.all()
            Starter = StarterSheet.objects.all()
        else:
            checksheets = CheckSheet.objects.filter(assigned_users=request.user)
            Starter = StarterSheet.objects.filter(assigned_users=request.user)

        selected_checksheet = (
            get_object_or_404(CheckSheet, id=checksheet_id) if checksheet_id else None
        )
        reasons = RejectReason.objects.all()
        if not selected_checksheet and checksheet_id:
            return HttpResponse("CheckSheet not found", status=404)
        images = selected_checksheet.images.all() if selected_checksheet else []
        zones = selected_checksheet.zones.all() if selected_checksheet else []

        today = timezone.now().date()

        # Get current time and determine current shift
        IST = pytz.timezone("Asia/Kolkata")
        current_time = timezone.now().astimezone(IST).time()
        try:
            shift_times = Shifttime.objects.first()
            current_shift = "None"
            if shift_times.shift_A_start <= current_time <= shift_times.shift_A_end:
                current_shift = "A"
            elif shift_times.shift_B_start <= current_time <= shift_times.shift_B_end:
                current_shift = "B"
        except Shifttime.DoesNotExist:
            current_shift = "None"

        # Get pending acknowledgments for today, current user, and current shift
        pending_acknowledgments_checksheet = FilledCheckSheet.objects.filter(
            user=request.user,
            timestamp__date=today,
            line=selected_checksheet.line,
            shift=current_shift,  # Filter by current shift
        )

        # Consolidated Data for the current shift
        consolidated_data = {}
        for entry in pending_acknowledgments_checksheet:
            key = (entry.checksheet.name, entry.shift)
            if key not in consolidated_data:
                consolidated_data[key] = defaultdict(int)
            for k, v in entry.status_data.items():
                if v == "Yes":
                    consolidated_data[key][k] += 1
                elif isinstance(v, int):
                    consolidated_data[key][k] += v
                elif k == "completely_reject":
                    if isinstance(v, int):
                        consolidated_data[key][k] += v
                    else:
                        consolidated_data[key]["completely_reject"] += 1

        # Prepare chart data for the current shift
        chart_labels = []
        chart_values = []
        for key, value_dict in consolidated_data.items():
            if key[1] == current_shift:  # Only include data for the current shift
                for zone_label, count in value_dict.items():
                    chart_labels.append(zone_label)
                    chart_values.append(count)

        if request.method == "POST":
            status_data = {}
            shift = request.POST.get("shift", current_shift)
            line = request.POST.get(
                "line", selected_checksheet.line if selected_checksheet else None
            )

            for zone in zones:
                user_input = request.POST.get(f"zone_{zone.id}", "").strip()
                if zone.input_type == "checkbox":
                    status_data[zone.name] = "Yes" if user_input else "No"
                elif zone.input_type == "int":
                    status_data[zone.name] = (
                        int(user_input) if user_input.isdigit() else 0
                    )
                elif zone.input_type == "float":
                    try:
                        status_data[zone.name] = (
                            float(user_input) if user_input else 0.0
                        )
                    except ValueError:
                        status_data[zone.name] = 0.0

            reject_reason = request.POST.get("reject_reason", "").strip()
            if reject_reason:
                status_data["completely_reject"] = reject_reason

            FilledCheckSheet.objects.create(
                checksheet=selected_checksheet,
                user=request.user,
                status_data=status_data,
                shift=shift,
                timestamp=timezone.now(),
                line=line,
            )

            if request.headers.get("X-Requested-With") == "XMLHttpRequest":
                return JsonResponse({"success": True})

            return redirect("fill_checksheet_detail", checksheet_id=checksheet_id)
        for i in reasons:
            print(i.reason)
        return render(
            request,
            "checksheet/fill_checksheet.html",
            {
                "checksheets": checksheets,
                "selected_checksheet": selected_checksheet,
                "zones": zones,
                "images": images,
                "Starter": Starter,
                "consolidated_data": consolidated_data,
                "chart_labels": chart_labels,
                "chart_values": chart_values,
                "current_shift": current_shift,
                "reasons":reasons,
            },
        )
    return render(request, "access_denied.html")


# ----------------------------------------- Create StarterSheet--------------------------------#

@login_required
def create_startersheet(request):
    if request.user.role == "admin" or has_page_access(
        request.user, "all_startersheet"
    ):
        all_users = get_user_model().objects.all()

        checksheets = (
            CheckSheet.objects.all()
            if request.user.role == "admin"
            else CheckSheet.objects.filter(assigned_users=request.user)
        )
        Starter = (
            StarterSheet.objects.all()
            if request.user.role == "admin"
            else StarterSheet.objects.filter(assigned_users=request.user)
        )

        if request.method == "POST":
            name = request.POST.get("name")
            line = request.POST.get("line")
            user_ids = request.POST.getlist("user_ids")  # For assigned users
            level_1_approver_id = request.POST.get("level_1_approver")
            level_2_approver_id = request.POST.get("level_2_approver")
            require_level_3 = request.POST.get("require_level_3") == "True"

            if name:
                # Create the StarterSheet
                startersheet = StarterSheet.objects.create(
                    name=name,
                    line=line,
                    created_by=request.user,
                    level_1_approver_id=level_1_approver_id if level_1_approver_id else None,
                    level_2_approver_id=level_2_approver_id if level_2_approver_id else None,
                    require_level_3_approval=require_level_3,
                )

                # Assign users
                if user_ids:
                    startersheet.assigned_users.set(user_ids)

                # Handle zones
                i = 0
                while f"zone_{i}" in request.POST:
                    zone_name = request.POST.get(f"zone_{i}")
                    zone_type = request.POST.get(f"zone_type_{i}")
                    min_value = request.POST.get(f"zone_min_{i}")
                    max_value = request.POST.get(f"zone_max_{i}")
                    unit = request.POST.get(f"zone_unit_{i}")
                    custom_unit = request.POST.get(f"zone_custom_unit_{i}")
                    final_unit = custom_unit if custom_unit else unit
                    check_method = request.POST.get(f"zone_check_method_{i}")
                    checkbox_label = request.POST.get(f"zone_checkbox_label_{i}")
                    zone_image = request.FILES.get(f"zone_image_{i}")
                    checkbox_default = request.POST.get(f"zone_default_{i}")

                    if zone_name and zone_type:
                        if zone_type == "checkbox":
                            min_value = "Yes" if checkbox_default else "No"
                            max_value = "Yes" if checkbox_default else "No"
                            final_unit = None
                        else:
                            min_value = min_value if min_value else None
                            max_value = max_value if max_value else None
                            checkbox_label = None

                        StarterZone.objects.create(
                            startersheet=startersheet,
                            name=zone_name,
                            type=zone_type,
                            min_value=min_value,
                            max_value=max_value,
                            unit=final_unit,
                            check_method=check_method,
                            image=zone_image,
                            standard=checkbox_label,
                        )
                    i += 1
                messages.success(
                    request,
                    "StarterSheet created successfully!",
                    extra_tags="StarterSheet_creation",
                )
                return redirect("all_startersheet")

        return render(
            request,
            "checksheet/create_startersheets.html",
            {
                "checksheets": checksheets,
                "Starter": Starter,
                "all_users": all_users,
            },
        )
    return render(request, "checksheet/access_denied.html")


# ----------------------------------------- All StarterSheet--------------------------------#


@login_required
def all_startersheet(request):
    if request.user.role == "admin" or has_page_access(
        request.user, "all_startersheet"
    ):
        # Admin sees all CheckSheets and StarterSheets
        checksheets = CheckSheet.objects.all()
        Starter = StarterSheet.objects.prefetch_related("assigned_users", "zones")
    else:
        # Operators see only assigned CheckSheets and StarterSheets
        checksheets = CheckSheet.objects.filter(assigned_users=request.user)
        Starter = StarterSheet.objects.filter(
            assigned_users=request.user
        ).prefetch_related("zones")
    all_users = get_user_model().objects.all()
    return render(
        request,
        "checksheet/all_startersheets.html",
        {
            "Starter": Starter,  # Using 'Starter' instead of 'startersheets'
            "checksheets": checksheets,
            "all_users": all_users,
        },
    )


# ----------------------------------------- Update StarterSheet-------------------------------#

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import StarterSheet, StarterZone

@login_required
def update_startersheet(request, startersheet_id):
    if request.user.role != "admin" and not has_page_access(request.user, "all_startersheet"):
        return render(request, "checksheet/access_denied.html")

    # Get the StarterSheet object
    startersheet = get_object_or_404(StarterSheet, id=startersheet_id)
    existing_zones = list(StarterZone.objects.filter(startersheet=startersheet))
    all_users = get_user_model().objects.all()  # Fetch all users for dropdowns

    if request.method == "POST":
        # Handle zone deletion
        delete_zone_id = request.POST.get("delete_zone")
        if delete_zone_id:
            try:
                zone = StarterZone.objects.get(id=delete_zone_id, startersheet=startersheet)
                zone.delete()
                messages.success(request, "Parameter deleted successfully!", extra_tags="zone_delete")
                return redirect("update_startersheet", startersheet_id=startersheet_id)
            except StarterZone.DoesNotExist:
                messages.error(request, "Parameter not found.", extra_tags="zone_delete_error")
                return redirect("update_startersheet", startersheet_id=startersheet_id)

        # Update startersheet details
        startersheet.name = request.POST.get("name", startersheet.name)
        startersheet.line = request.POST.get("line", startersheet.line)

        # Handle assigned users
        user_ids = request.POST.getlist("user_ids")  # Get list of selected user IDs
        if user_ids:
            startersheet.assigned_users.set(user_ids)  # Update many-to-many relationship
        else:
            startersheet.assigned_users.clear()  # Clear if no users selected

        # Handle Level 1 Approver
        level_1_approver_id = request.POST.get("level_1_approver")
        startersheet.level_1_approver = get_user_model().objects.get(id=level_1_approver_id) if level_1_approver_id else None

        # Handle Level 2 Approver
        level_2_approver_id = request.POST.get("level_2_approver")
        startersheet.level_2_approver = get_user_model().objects.get(id=level_2_approver_id) if level_2_approver_id else None

        # Handle Level 3 Approval
        require_level_3 = request.POST.get("require_level_3") == "True"
        startersheet.require_level_3_approval = require_level_3

        startersheet.save()

        updated_zone_ids = set()
        new_zones = []

        # Process existing zones
        for i, zone in enumerate(existing_zones):
            zone_name = request.POST.get(f"zone_{i}")
            if zone_name is None:  # Zone was removed
                continue
            zone_type = request.POST.get(f"zone_type_{i}", zone.type)
            checkbox_label = request.POST.get(f"zone_checkbox_label_{i}")
            zone_unit = request.POST.get(f"zone_unit_{i}", "")
            zone_custom_unit = request.POST.get(f"zone_custom_unit_{i}", "")
            unit = zone_custom_unit if zone_custom_unit else zone_unit
            check_method = request.POST.get(f"zone_check_method_{i}", "")
            zone_image = request.FILES.get(f"zone_image_{i}")

            if zone_type == "checkbox":
                checkbox_value = request.POST.get(f"zone_default_{i}")
                min_value = "Yes" if checkbox_value == "on" else "No"
                max_value = "Yes" if checkbox_value == "on" else "No"
                zone.unit = None
                zone.standard = checkbox_label
            else:
                min_value = request.POST.get(f"zone_min_{i}")
                max_value = request.POST.get(f"zone_max_{i}")
                min_value = int(min_value) if min_value and min_value.isdigit() else zone.min_value
                max_value = int(max_value) if max_value and max_value.isdigit() else zone.max_value
                zone.unit = unit
                zone.standard = None

            zone.name = zone_name
            zone.type = zone_type
            zone.min_value = min_value
            zone.max_value = max_value
            zone.check_method = check_method
            if zone_image:
                zone.image = zone_image
            zone.save()
            updated_zone_ids.add(zone.id)

        # Process new zones
        i = 1
        while f"zone_new_{i}" in request.POST:
            zone_name = request.POST.get(f"zone_new_{i}")
            zone_type = request.POST.get(f"zone_type_new_{i}")
            checkbox_label = request.POST.get(f"zone_checkbox_label_new_{i}")
            zone_unit = request.POST.get(f"zone_unit_new_{i}", "")
            zone_custom_unit = request.POST.get(f"zone_custom_unit_new_{i}", "")
            unit = zone_custom_unit if zone_custom_unit else zone_unit
            check_method = request.POST.get(f"zone_check_method_new_{i}", "")
            zone_image = request.FILES.get(f"zone_image_new_{i}")

            if zone_type == "checkbox":
                checkbox_value = request.POST.get(f"zone_default_new_{i}")
                min_value = "Yes" if checkbox_value == "on" else "No"
                max_value = "Yes" if checkbox_value == "on" else "No"
                unit = None
                standard = checkbox_label
            else:
                min_value = request.POST.get(f"zone_min_new_{i}")
                max_value = request.POST.get(f"zone_max_new_{i}")
                min_value = int(min_value) if min_value and min_value.isdigit() else None
                max_value = int(max_value) if max_value and max_value.isdigit() else None
                standard = None

            # Create new zone
            new_zone = StarterZone.objects.create(
                startersheet=startersheet,
                name=zone_name,
                type=zone_type,
                min_value=min_value,
                max_value=max_value,
                unit=unit,
                standard=standard,
                check_method=check_method,
                image=zone_image
            )
            new_zones.append(new_zone)
            i += 1

        messages.success(request, "StarterSheet updated successfully!", extra_tags="StarterSheet_zone")
        return redirect("all_startersheet")

    # GET request - render the form
    context = {
        'startersheet': startersheet,
        'zones': existing_zones,
        'all_users': all_users,  # Add all_users to context
        'success': any(msg.tags == "StarterSheet_zone" for msg in messages.get_messages(request)),
    }
    return render(request, 'checksheet/update_startersheet.html', context)
# ----------------------------------------- Add StarterSheet Parameter--------------------------------#


# @login_required
# def Add_start_zone(request, startersheet_id):
#     if request.user.role != "admin" and not has_page_access(
#         request.user, "all_startersheet"
#     ):
#         return render(request, "checksheet/access_denied.html")

#     checksheets = (
#         CheckSheet.objects.all()
#         if request.user.role == "admin"
#         else CheckSheet.objects.filter(assigned_users=request.user)
#     )
#     Starter = (
#         StarterSheet.objects.all()
#         if request.user.role == "admin"
#         else StarterSheet.objects.filter(assigned_users=request.user)
#     )

#     startersheet = get_object_or_404(StarterSheet, id=startersheet_id)

#     if request.method == "POST":
#         # Get the new zone details from the form
#         zone_name = request.POST.get("zone_name")
#         zone_type = request.POST.get("zone_type")
#         zone_image = request.FILES.get("zone_image")

#         # Get the new field values
#         zone_unit = request.POST.get("zone_unit")
#         zone_custom_unit = request.POST.get("zone_custom_unit")
#         zone_check_method = request.POST.get("zone_check_method")
#         checkbox_label = request.POST.get("checkbox_text")

#         # Determine which unit to use (standard or custom)
#         unit = zone_custom_unit if zone_custom_unit else zone_unit

#         # Handle values based on zone type
#         if zone_type == "checkbox":
#             # Get checkbox state (on if checked, None if unchecked)
#             checkbox_value = request.POST.get("zone_default")
#             # Set min and max to "Yes" if checked, "No" if unchecked
#             min_value = "Yes" if checkbox_value == "on" else "No"
#             max_value = "Yes" if checkbox_value == "on" else "No"
#             unit = None
#             standard = checkbox_label
#         else:
#             # Handle numeric inputs for non-checkbox types
#             min_value = request.POST.get("zone_min")
#             max_value = request.POST.get("zone_max")
#             min_value = int(min_value) if min_value and min_value.isdigit() else None
#             max_value = int(max_value) if max_value and max_value.isdigit() else None
#             unit = unit
#             standard = None

#         if zone_name and zone_type:
#             # Create a new StarterZone object and associate it with the startersheet
#             StarterZone.objects.create(
#                 startersheet=startersheet,
#                 name=zone_name,
#                 type=zone_type,
#                 image=zone_image,
#                 min_value=min_value,
#                 max_value=max_value,
#                 unit=unit,  # Add the unit field
#                 check_method=zone_check_method,
#                 standard=standard,  # Add the check method field
#             )
#             messages.success(
#                 request,
#                 "Zone Added sucessfully!",
#                 extra_tags="StarterSheet_zone",
#             )
#             return redirect("all_startersheet")

#     return render(
#         request,
#         "checksheet/add_startzone.html",
#         {
#             "startersheet": startersheet,
#             "checksheets": checksheets,
#             "Starter": Starter,
#         },
#     )


# -----------------------------------------Fill StarterSheet--------------------------------#


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.utils import timezone
from .models import CheckSheet, StarterSheet, FilledStarterSheet, Shifttime
import json
import pytz

@login_required
def fill_starter_sheet(request, startersheet_id=None):
    if request.user.role in ["operator", "admin"] or has_page_access(request.user, "fill_starter_sheet"):
        if request.user.role == "admin":
            checksheets = CheckSheet.objects.all()
            Starter = StarterSheet.objects.all()
        else:
            checksheets = CheckSheet.objects.filter(assigned_users=request.user)
            Starter = StarterSheet.objects.filter(assigned_users=request.user)

        selected_startersheet = get_object_or_404(StarterSheet, id=startersheet_id) if startersheet_id else None
        zones = selected_startersheet.zones.all() if selected_startersheet else []
        today = timezone.now().date()

        IST = pytz.timezone("Asia/Kolkata")
        current_time = timezone.now().astimezone(IST).time()
        try:
            shift_times = Shifttime.objects.first()
            current_shift = "None"
            if shift_times.shift_A_start <= current_time <= shift_times.shift_A_end:
                current_shift = "A"
            elif shift_times.shift_B_start <= current_time <= shift_times.shift_B_end:
                current_shift = "B"
        except Shifttime.DoesNotExist:
            current_shift = "None"

        if request.method == "POST":
            try:
                data = json.loads(request.body)
                shift = data.get("shift")
                user = request.user
                line = data.get("line", selected_startersheet.line if selected_startersheet else None)
                out_of_range_reason = data.get("out_of_range_reason")

                existing_entry = FilledStarterSheet.objects.filter(
                    startersheet=selected_startersheet,
                    filled_by=user,
                    shift=shift,
                    line=line,
                    timestamp__date=today,
                ).exists()

                if existing_entry:
                    return JsonResponse(
                        {"error": "Data already filled for this user, shift, line, and sheet today"},
                        status=400,
                    )

                status_data = {}
                for zone_data in data["zones"]:
                    zone_id = zone_data["id"]
                    user_input = zone_data["value"]
                    zone = next((z for z in zones if str(z.id) == zone_id), None)
                    if not zone:
                        continue

                    zone_name = zone.name
                    if zone.type == "checkbox":
                        status_data[zone_name] = "Yes" if user_input == "Yes" else "No"
                    elif zone.type == "int":
                        status_data[zone_name] = int(user_input) if user_input.isdigit() else 0
                    elif zone.type == "float":
                        try:
                            float(user_input)
                            status_data[zone_name] = user_input
                        except ValueError:
                            status_data[zone_name] = "0.0"

                FilledStarterSheet.objects.create(
                    startersheet=selected_startersheet,
                    filled_by=user,
                    status_data=status_data,
                    shift=shift,
                    line=line,
                    out_of_range_reason=out_of_range_reason
                )

                return JsonResponse({"message": "Data saved successfully"}, status=200)
            except json.JSONDecodeError:
                return JsonResponse({"error": "Invalid data"}, status=400)

        return render(
            request,
            "checksheet/fill_starter_sheet.html",
            {
                "Starter": Starter,
                "selected_startersheet": selected_startersheet,
                "zones": zones,
                "checksheets": checksheets,
                "current_shift": current_shift,
            },
        )
    return render(request, "checksheet/access_denied.html")


# ----------------------------------------- Create User --------------------------------#
@login_required
@user_passes_test(lambda u: u.is_superuser)
def create_user(request):
    if request.user.role == "admin":
        checksheets = CheckSheet.objects.all()
        Starter = StarterSheet.objects.all()
    else:
        checksheets = CheckSheet.objects.filter(assigned_users=request.user)
        Starter = StarterSheet.objects.filter(assigned_users=request.user)
    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        username = request.POST.get("username")
        password = request.POST.get("password")
        email = request.POST.get("email")
        role = request.POST.get("role")
        phone_number = request.POST.get("phone_number")  # Get phone number from form

        if CustomUser.objects.filter(employee_id=employee_id).exists():
            messages.error(
                request,
                f"Username '{employee_id}' already exists. Please choose a different username.",
                extra_tags="user_exist",
            )
            return render(
                request, "checksheet/create_user.html"
            )  # Stay on the same page

        # Create and save user if username doesn't exist
        new_user = CustomUser.objects.create_user(
            username=username,
            email=email,
            password=password,
            employee_id=employee_id,
            role=role,
            phone_number=phone_number,  # Add phone number to the user object
        )
        new_user.save()
        messages.success(
            request,
            f"User '{username}' created successfully!",
            extra_tags="user_creation",
        )
        return redirect("user_list")  # Redirect to user list after success

    return render(
        request,
        "checksheet/create_user.html",
        {"checksheets": checksheets, "Starter": Starter},
    )


# ----------------------------------------- User List(all_user)--------------------------------#


@login_required
@user_passes_test(lambda u: u.is_superuser)
def user_list(request):
    if request.user.role == "admin":
        checksheets = CheckSheet.objects.all()
        Starter = StarterSheet.objects.all()
    else:
        checksheets = CheckSheet.objects.filter(assigned_users=request.user)
        Starter = StarterSheet.objects.filter(assigned_users=request.user)

    users = CustomUser.objects.filter(is_superuser=False)  # Exclude superusers

    return render(
        request,
        "checksheet/user_list.html",
        {"users": users, "checksheets": checksheets, "Starter": Starter},
    )


# ----------------------------------------- Edit Users--------------------------------#


@login_required
@user_passes_test(lambda u: u.is_superuser)
def edit_user(request, user_id):
    if request.user.role == "admin":
        checksheets = CheckSheet.objects.all()
        Starter = StarterSheet.objects.all()
    else:
        checksheets = CheckSheet.objects.filter(assigned_users=request.user)
        Starter = StarterSheet.objects.filter(assigned_users=request.user)

    user = get_object_or_404(CustomUser, id=user_id)

    if request.method == "POST":
        user.username = request.POST.get("username")
        user.email = request.POST.get("email")
        user.role = request.POST.get("role")
        user.phone_number = request.POST.get("phone_number")
        # Update is_active field
        user.is_active = request.POST.get("is_active") == "on"  # Checkbox returns "on" if checked

        # Handle password change if provided
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if password and password == confirm_password:
            user.set_password(password)

        user.save()
        return redirect("user_list")

    return render(
        request,
        "checksheet/edit_user.html",
        {"user": user, "checksheets": checksheets, "Starter": Starter},
    )


# ----------Dashboard function for (operator shift_incharge and  quality_incharge----------------------#


@login_required
@user_passes_test(is_quality_incharge)
def quality_incharge_dashboard(request):

    if request.user.role == "admin":
        # Admins can see all pending starter sheets
        pending_acknowledgments = FilledStarterSheet.objects.filter(
            Q(
                approval_status__in=[
                    "pending",
                    "level_1_approved",
                    "level_2_approved",
                ],
                requires_level_3_approval=True,
            )
        ).exclude(approval_status="rejected")

        # For admins, count starter sheets they specifically approved (excluding rejected ones)
        approved_count = (
            FilledStarterSheet.objects.filter(
                Q(level_1_approval_id=request.user.id)
                | Q(level_2_approval_id=request.user.id)
                | Q(level_3_approval_id=request.user.id)
            )
            .exclude(approval_status="rejected")
            .count()
        )

    else:
        # Regular users see starter sheets they are assigned to approve
        pending_acknowledgments = FilledStarterSheet.objects.filter(
            # Sheets pending their approval
            Q(assigned_level_1_approver=request.user, approval_status="pending")
            | Q(
                assigned_level_2_approver=request.user,
                approval_status="level_1_approved",
            )
        ).exclude(approval_status="rejected")

        # Count starter sheets approved by this user (excluding rejected ones)
        approved_count = (
            FilledStarterSheet.objects.filter(
                Q(
                    level_1_approval_id=request.user.id,
                    approval_status__in=[
                        "level_1_approved",
                        "level_2_approved",
                        "completed",
                    ],
                )
                | Q(
                    level_2_approval_id=request.user.id,
                    approval_status__in=["level_2_approved", "completed"],
                )
            )
            .exclude(approval_status="rejected")
            .count()
        )

    # Get the count of pending starter sheets
    pending_count = pending_acknowledgments.count()

    # Convert to values for template
    pending_acknowledgments = pending_acknowledgments.values(
        "id",
        "startersheet__id",
        "startersheet__name",
        "filled_by__id",
        "filled_by__username",
        "shift",
        "timestamp",
        "approval_status",
        "rejection_reason",
        "rejected_by_id",
        "level_1_approval_id",
        "level_2_approval_id",
    ).order_by("timestamp")

    # Now handle FilledCheckSheet data
    if request.user.role == "admin":
        # Admins can see all pending check sheets that require approval AND are flagged for acknowledgment
        pending_check_acknowledgments = FilledCheckSheet.objects.filter(
            Q(send_acknowledgment=True)
            & (
                Q(
                    approval_status__in=[
                        "pending",
                        "level_1_approved",
                        "level_2_approved",
                    ],
                    requires_level_3_approval=True,
                )
                |
                # Show rejected sheets to admin as well
                Q(approval_status="rejected")
            )
        )

        # For admins, get check sheets they specifically approved (excluding rejected ones)
        approved_check_sheets = FilledCheckSheet.objects.filter(
            Q(send_acknowledgment=True)
            & (
                Q(level_1_approval_id=request.user.id)
                | Q(level_2_approval_id=request.user.id)
                | Q(level_3_approval_id=request.user.id)
            )
        ).exclude(approval_status="rejected")
    else:
        # Regular users see check sheets they are assigned to approve
        pending_check_acknowledgments = FilledCheckSheet.objects.filter(
            Q(send_acknowledgment=True)
            & (
                Q(assigned_level_1_approver=request.user, approval_status="pending")
                | Q(
                    assigned_level_2_approver=request.user,
                    approval_status="level_1_approved",
                )
            )
        ).exclude(approval_status="rejected")

        # Get check sheets approved by this user (excluding rejected ones)
        approved_check_sheets = FilledCheckSheet.objects.filter(
            Q(send_acknowledgment=True)
            & (
                Q(
                    level_1_approval_id=request.user.id,
                    approval_status__in=[
                        "level_1_approved",
                        "level_2_approved",
                        "completed",
                    ],
                )
                | Q(
                    level_2_approval_id=request.user.id,
                    approval_status__in=["level_2_approved", "completed"],
                )
            )
        ).exclude(approval_status="rejected")

    # Count unique combinations for pending check sheets
    pending_unique_combinations = set()
    for check in pending_check_acknowledgments:
        # Extract date from timestamp
        date = check.timestamp.date()
        # Create a unique identifier tuple
        unique_key = (date, check.shift, check.user_id, check.checksheet_id)
        pending_unique_combinations.add(unique_key)

    # Count unique combinations for approved check sheets
    approved_unique_combinations = set()
    for check in approved_check_sheets:
        # Extract date from timestamp
        date = check.timestamp.date()
        # Create a unique identifier tuple
        unique_key = (date, check.shift, check.user_id, check.checksheet_id)
        approved_unique_combinations.add(unique_key)

    # Get the consolidated counts
    check_pending_count = len(pending_unique_combinations)
    check_approved_count = len(approved_unique_combinations)

    print(
        f"Consolidated counts: {check_approved_count} approved, {check_pending_count} pending"
    )

    return render(
        request,
        "checksheet/quality_incharge_dashboard.html",
        {
            "pending_acknowledgments": pending_acknowledgments,
            "startersheet_pending_count": pending_count,
            "startersheet_acknowledged_count": approved_count,
            "checksheet_pending_count": check_pending_count,
            "checksheet_acknowledged_count": check_approved_count,
        },
    )


@login_required
@user_passes_test(is_shift_incharge)
def shift_incharge_dashboard(request):
    # Get accepted form requests
    form_requests = FormRequest.objects.filter(status="Accepted")

    # Get pending and approved counts for FilledStarterSheet
    if request.user.role == "admin":
        # Admins can see all pending starter sheets
        pending_acknowledgments = FilledStarterSheet.objects.filter(
            Q(
                approval_status__in=[
                    "pending",
                    "level_1_approved",
                    "level_2_approved",
                ],
                requires_level_3_approval=True,
            )
        ).exclude(approval_status="rejected")

        # For admins, count starter sheets they specifically approved (excluding rejected ones)
        approved_count = (
            FilledStarterSheet.objects.filter(
                Q(level_1_approval_id=request.user.id)
                | Q(level_2_approval_id=request.user.id)
                | Q(level_3_approval_id=request.user.id)
            )
            .exclude(approval_status="rejected")
            .count()
        )

    else:
        # Regular users see starter sheets they are assigned to approve
        pending_acknowledgments = FilledStarterSheet.objects.filter(
            # Sheets pending their approval
            Q(assigned_level_1_approver=request.user, approval_status="pending")
            | Q(
                assigned_level_2_approver=request.user,
                approval_status="level_1_approved",
            )
        ).exclude(approval_status="rejected")

        # Count starter sheets approved by this user (excluding rejected ones)
        approved_count = (
            FilledStarterSheet.objects.filter(
                Q(
                    level_1_approval_id=request.user.id,
                    approval_status__in=[
                        "level_1_approved",
                        "level_2_approved",
                        "completed",
                    ],
                )
                | Q(
                    level_2_approval_id=request.user.id,
                    approval_status__in=["level_2_approved", "completed"],
                )
            )
            .exclude(approval_status="rejected")
            .count()
        )

    # Get the count of pending starter sheets
    pending_count = pending_acknowledgments.count()

    # Convert to values for template
    pending_acknowledgments = pending_acknowledgments.values(
        "id",
        "startersheet__id",
        "startersheet__name",
        "filled_by__id",
        "filled_by__username",
        "shift",
        "timestamp",
        "approval_status",
        "rejection_reason",
        "rejected_by_id",
        "level_1_approval_id",
        "level_2_approval_id",
    ).order_by("timestamp")

    # Now handle FilledCheckSheet data
    if request.user.role == "admin":
        # Admins can see all pending check sheets that require approval AND are flagged for acknowledgment
        pending_check_acknowledgments = FilledCheckSheet.objects.filter(
            Q(send_acknowledgment=True)
            & (
                Q(
                    approval_status__in=[
                        "pending",
                        "level_1_approved",
                        "level_2_approved",
                    ],
                    requires_level_3_approval=True,
                )
                |
                # Show rejected sheets to admin as well
                Q(approval_status="rejected")
            )
        )

        # For admins, get check sheets they specifically approved (excluding rejected ones)
        approved_check_sheets = FilledCheckSheet.objects.filter(
            Q(send_acknowledgment=True)
            & (
                Q(level_1_approval_id=request.user.id)
                | Q(level_2_approval_id=request.user.id)
                | Q(level_3_approval_id=request.user.id)
            )
        ).exclude(approval_status="rejected")
    else:
        # Regular users see check sheets they are assigned to approve
        pending_check_acknowledgments = FilledCheckSheet.objects.filter(
            Q(send_acknowledgment=True)
            & (
                Q(assigned_level_1_approver=request.user, approval_status="pending")
                | Q(
                    assigned_level_2_approver=request.user,
                    approval_status="level_1_approved",
                )
            )
        ).exclude(approval_status="rejected")

        # Get check sheets approved by this user (excluding rejected ones)
        approved_check_sheets = FilledCheckSheet.objects.filter(
            Q(send_acknowledgment=True)
            & (
                Q(
                    level_1_approval_id=request.user.id,
                    approval_status__in=[
                        "level_1_approved",
                        "level_2_approved",
                        "completed",
                    ],
                )
                | Q(
                    level_2_approval_id=request.user.id,
                    approval_status__in=["level_2_approved", "completed"],
                )
            )
        ).exclude(approval_status="rejected")

    # Count unique combinations for pending check sheets
    pending_unique_combinations = set()
    for check in pending_check_acknowledgments:
        # Extract date from timestamp
        date = check.timestamp.date()
        # Create a unique identifier tuple
        unique_key = (date, check.shift, check.user_id, check.checksheet_id)
        pending_unique_combinations.add(unique_key)

    # Count unique combinations for approved check sheets
    approved_unique_combinations = set()
    for check in approved_check_sheets:
        # Extract date from timestamp
        date = check.timestamp.date()
        # Create a unique identifier tuple
        unique_key = (date, check.shift, check.user_id, check.checksheet_id)
        approved_unique_combinations.add(unique_key)

    # Get the consolidated counts
    check_pending_count = len(pending_unique_combinations)
    check_approved_count = len(approved_unique_combinations)

    print(
        f"Consolidated counts: {check_approved_count} approved, {check_pending_count} pending"
    )

    # Return the data to the template with both starter sheets and check sheets
    return render(
        request,
        "checksheet/shift_incharge_dashboard.html",
        {
            "form_requests": form_requests,
            "pending_acknowledgments": pending_acknowledgments,
            "startersheet_pending_count": pending_count,
            "startersheet_acknowledged_count": approved_count,
            "checksheet_pending_count": check_pending_count,
            "checksheet_acknowledged_count": check_approved_count,
        },
    )


@login_required
def operator_dashboard(request):
    if request.user.role in ["operator","admin"]:
        if request.user.role == "admin":
            checksheets = CheckSheet.objects.all()
            Starter = StarterSheet.objects.all()
        else:
            checksheets = CheckSheet.objects.filter(assigned_users=request.user)
            Starter = StarterSheet.objects.filter(assigned_users=request.user)

        current_date = now().date()  # ✅ Get today's date (ignoring time)
        user = request.user

        # Get current time and determine current shift
        IST = pytz.timezone("Asia/Kolkata")
        current_time = now().astimezone(IST).time()
        # Get current time in IST
        try:
            shift_times = (
                Shifttime.objects.first()
            )  # Get the first record from Shifttime model
            current_shift = "None"  # Default value

            # Check if current time is in shift A
            if shift_times.shift_A_start <= current_time <= shift_times.shift_A_end:
                current_shift = "A"
            # Check if current time is in shift B
            elif shift_times.shift_B_start <= current_time <= shift_times.shift_B_end:
                current_shift = "B"
        except Shifttime.DoesNotExist:
            current_shift = "None"

        # ✅ Get assigned POCs

        # ✅ Check if the user has filled at least one Starter Sheet today (ignoring time)
        has_filled_starter_sheet_today = FilledStarterSheet.objects.filter(
            filled_by=user, timestamp__date=current_date, shift=current_shift
        ).exists()

        return render(
            request,
            "checksheet/operator_dashboard.html",
            {
                "checksheets": checksheets,
                "Starter": Starter,
                "has_filled_starter_sheet": has_filled_starter_sheet_today,
            },
        )
    return render(request, "checksheet/access_denied.html")


# ----------------------------------------- Upload POC--------------------------------#
@login_required
def upload_poc(request):
    if request.user.role == "admin" or has_page_access(request.user, "upload_poc"):
        startersheets = StarterSheet.objects.all()
        poc_files = list(
            POCUpload.objects.prefetch_related("assigned_startersheets").all()[:15]
        )  # Get latest 15 POCs

        if request.method == "POST" and "save_poc" in request.POST:
            # Update existing POCs
            for poc in poc_files:
                file_key = f"pdf_{poc.id}"
                if file_key in request.FILES:
                    poc.pdf = request.FILES[file_key]
                    poc.save()

                # Assign StarterSheets to existing POCs
                selected_startersheets = request.POST.getlist(
                    f"poc_{poc.id}_startersheets"
                )
                poc.assigned_startersheets.set(selected_startersheets)

            # Handle new POC uploads
            for i in range(15 - len(poc_files)):  # Remaining slots for new uploads
                file_key = f"new_pdf_{i}"
                if file_key in request.FILES:
                    uploaded_file = request.FILES[file_key]
                    new_poc = POCUpload.objects.create(pdf=uploaded_file)

                    # Assign StarterSheets to new POCs
                    selected_startersheets = request.POST.getlist(
                        f"new_poc_{i}_startersheets"
                    )
                    new_poc.assigned_startersheets.set(selected_startersheets)
                    poc_files.append(new_poc)
            messages.success(
                request,
                "ops updated sucessfully!",
                extra_tags="ops",
            )
            return redirect("upload_poc")  # Redirect after saving

        remaining_inputs = 15 - len(poc_files)

        return render(
            request,
            "checksheet/upload_poc.html",
            {
                "poc_files": poc_files,
                "remaining_inputs": range(remaining_inputs),
                "startersheets": startersheets,
            },
        )

    return render(request, "checksheet/access_denied.html")


# ----------------------------------------- View POC--------------------------------#
@login_required
def view_poc(request):
    if request.user.role == "admin":
        # Admin sees all CheckSheets, StarterSheets, and POC PDFs
        checksheets = CheckSheet.objects.all()
        Starter = StarterSheet.objects.all()
        poc_pdfs = POCUpload.objects.all()
    else:
        # Operators see only assigned CheckSheets and StarterSheets
        checksheets = CheckSheet.objects.filter(assigned_users=request.user)
        Starter = StarterSheet.objects.filter(assigned_users=request.user)

        # Get POC PDFs related to the assigned CheckSheets
        poc_pdfs = POCUpload.objects.filter(
            assigned_startersheets__in=Starter
        ).distinct()
    poc_read_status = {
        poc.id: POCReadStatus.objects.filter(user=request.user, poc=poc).exists()
        for poc in poc_pdfs
    }

    return render(
        request,
        "checksheet/view_poc.html",
        {
            "poc_pdfs": poc_pdfs,
            "checksheets": checksheets,
            "Starter": Starter,
            "poc_read_status": poc_read_status,
        },
    )


# ----------------------------------------- Acknowledgment--------------------------------#
@login_required
def acknowledgment_list(request):
    if request.user.role == "admin" or has_page_access(
        request.user, "acknowledgment_list"
    ):
        checksheets = CheckSheet.objects.all()
        Starter = StarterSheet.objects.all()
        is_admin = request.user.role == "admin"
        pending_requests = (
            FormRequest.objects.filter(status="Pending") if is_admin else None
        )
        if request.user.role == "admin":
            # Admins can see all pending sheets
            pending_acknowledgments = FilledStarterSheet.objects.filter(
                Q(
                    approval_status__in=[
                        "level_2_approved",
                    ],
                    requires_level_3_approval=True,
                )
            )
        else:

            pending_acknowledgments = FilledStarterSheet.objects.filter(
                # Sheets pending their approval
                Q(assigned_level_1_approver=request.user, approval_status="pending")
                | Q(
                    assigned_level_2_approver=request.user,
                    approval_status="level_1_approved",
                )
                | Q(
                    rejected_by_id=1,
                    assigned_level_2_approver=request.user,
                    approval_status="rejected",
                )
                |
                # If rejected by level 2, show to level 1
                Q(
                    level_1_approval_id=request.user.id,
                    rejected_by_id=F("assigned_level_2_approver"),
                    approval_status="rejected",
                )
                |
                # If rejected by level 3, show to level 2
                Q(
                    level_2_approval_id=request.user.id,
                    rejected_by_id=F("level_3_approval_id"),
                    approval_status="rejected",
                )
            )

        # Convert to values for template
        pending_acknowledgments = pending_acknowledgments.values(
            "id",
            "startersheet__id",
            "startersheet__name",
            "filled_by__id",
            "filled_by__username",
            "shift",
            "timestamp",
            "approval_status",
            "rejection_reason",  # Add this
            "rejected_by_id",  # Add this
            "level_1_approval_id",
            "level_2_approval_id",
            "line",
        ).order_by("timestamp")

        # For admins
        if request.user.role == "admin":
            # Admins can see all pending check sheets that require approval AND are flagged for acknowledgment
            pending_check_acknowledgments = FilledCheckSheet.objects.filter(
                Q(send_acknowledgment=True)  # Changed from OR to AND using &
                & (
                    Q(
                        approval_status__in=[
                            "level_2_approved",
                        ],
                        requires_level_3_approval=True,
                    )
                )
            )
        else:
            # Regular users see check sheets they are assigned to approve or that they approved but were rejected
            # AND that are flagged for acknowledgment
            pending_check_acknowledgments = FilledCheckSheet.objects.filter(
                Q(send_acknowledgment=True)  # Changed from OR to AND using &
                & (
                    Q(assigned_level_1_approver=request.user, approval_status="pending")
                    | Q(
                        assigned_level_2_approver=request.user,
                        approval_status="level_1_approved",
                    )
                    |
                    # Show rejected sheets based on appropriate flow:
                    # If rejected by admin (user_id=1), show only to level 2 approver
                    Q(
                        rejected_by_id=1,
                        assigned_level_2_approver=request.user,
                        approval_status="rejected",
                    )
                    |
                    # If rejected by level 2, show to level 1
                    Q(
                        level_1_approval_id=request.user.id,
                        rejected_by_id=F("assigned_level_2_approver"),
                        approval_status="rejected",
                    )
                    |
                    # If rejected by level 3, show to level 2
                    Q(
                        level_2_approval_id=request.user.id,
                        rejected_by_id=F("level_3_approval_id"),
                        approval_status="rejected",
                    )
                )
            )

        # Process and consolidate check sheet data
        consolidated_data = {}
        for entry in pending_check_acknowledgments:
            # Convert timestamp to date string for grouping
            entry_date = entry.timestamp.date()
            # Add line and employee_id to the key tuple
            key = (
                entry.checksheet.name,
                entry.user.username,
                entry.shift,
                entry_date,
                entry.line,
                entry.user.employee_id,
            )
            if key not in consolidated_data:
                consolidated_data[key] = defaultdict(int)
                consolidated_data[key]["reject_reasons"] = set()

            for k, v in entry.status_data.items():
                if v == "Yes":
                    consolidated_data[key][k] += 1
                elif isinstance(v, int):
                    consolidated_data[key][k] += v
                elif k == "completely_reject":
                    if isinstance(v, int):
                        consolidated_data[key][k] += v
                    else:
                        consolidated_data[key]["completely_reject"] += 1

        # Convert consolidated check sheet data to structured list
        consolidated_check_entries = []
        for (
            checksheet_name,
            username,
            shift,
            entry_date,
            line,
            employee_id,  # Added employee_id to the unpacking
        ), status_counts in consolidated_data.items():
            status_display = ", ".join(
                f"{k}: {v}" for k, v in status_counts.items() if k != "reject_reasons"
            )
            reject_reasons = (
                ", ".join(status_counts["reject_reasons"])
                if status_counts["reject_reasons"]
                else ""
            )

            if reject_reasons:
                status_display += f", completely_reject: {reject_reasons}"

            consolidated_check_entries.append(
                {
                    "checksheet_name": checksheet_name,
                    "username": username,
                    "shift": shift,
                    "timestamp": entry_date,
                    "date": entry_date.strftime("%Y-%m-%d"),
                    "line": line,
                    "employee_id": employee_id,  # Add employee_id to the output dictionary
                    "acknowledgment": entry.approval_status,
                    "status_data": status_display,
                }
            )
        print(consolidated_check_entries, "cffffff")
        return render(
            request,
            "checksheet/acknowledge.html",
            {
                "checksheets": checksheets,
                "Starter": Starter,
                "pending_acknowledgments_checksheet": consolidated_check_entries,
                "pending_acknowledgments": pending_acknowledgments,
                "pending_requests": pending_requests,
            },
        )
    return render(request, "checksheet/access_denied.html")


# ----------------------------------------- view filled startersheet --------------------------------#


def view_filled_startersheet(request, startersheet_id, user_id, shift, id):
    if request.user.role == "admin" or has_page_access(
        request.user, "acknowledgment_list"
    ):
        filled_entries = FilledStarterSheet.objects.filter(
            startersheet_id=startersheet_id, filled_by_id=user_id, shift=shift, id=id
        )

        if not filled_entries:
            return redirect("acknowledgment_list")

        entry = filled_entries[0]
        if isinstance(entry.status_data, str):
            entry.status_data = json.loads(entry.status_data)

        # Extract sheet name
        sheet_name = entry.startersheet if entry else None

        # Fetch min and max values for each parameter
        parameter_thresholds = {
            zone.name: {"min": zone.min_value, "max": zone.max_value}
            for zone in StarterZone.objects.filter(startersheet_id=startersheet_id)
        }

        # Get all users involved in the approval process
        approval_hierarchy = {}

        # Level 1 approver info
        if entry.assigned_level_1_approver_id:
            try:
                level1_approver = User.objects.get(
                    id=entry.assigned_level_1_approver_id
                )
                approval_hierarchy["level1"] = {
                    "assigned_to": level1_approver.username,
                    "status": "Pending",
                    "timestamp": None,
                }

                # Check if approved by level 1
                if entry.level_1_approval_id:
                    level1_approver_action = User.objects.get(
                        id=entry.level_1_approval_id
                    )
                    approval_hierarchy["level1"].update(
                        {
                            "status": "Approved",
                            "action_by": level1_approver_action.username,
                            "timestamp": entry.level_1_approval_timestamp,
                        }
                    )
            except User.DoesNotExist:
                approval_hierarchy["level1"] = {
                    "assigned_to": "Unknown User",
                    "status": "Unknown",
                }

        # Level 2 approver info
        if entry.assigned_level_2_approver_id:
            try:
                level2_approver = User.objects.get(
                    id=entry.assigned_level_2_approver_id
                )
                approval_hierarchy["level2"] = {
                    "assigned_to": level2_approver.username,
                    "status": "Pending",
                    "timestamp": None,
                }

                # Check if approved by level 2
                if entry.level_2_approval_id:
                    level2_approver_action = User.objects.get(
                        id=entry.level_2_approval_id
                    )
                    approval_hierarchy["level2"].update(
                        {
                            "status": "Approved",
                            "action_by": level2_approver_action.username,
                            "timestamp": entry.level_2_approval_timestamp,
                        }
                    )
            except User.DoesNotExist:
                approval_hierarchy["level2"] = {
                    "assigned_to": "Unknown User",
                    "status": "Unknown",
                }

        # Level 3 approver info (if required)
        if entry.requires_level_3_approval:
            approval_hierarchy["level3"] = {
                "assigned_to": "Admin",
                "status": "Pending",
                "timestamp": None,
            }

            # Check if approved by level 3
            if entry.level_3_approval_id:
                try:
                    level3_approver_action = User.objects.get(
                        id=entry.level_3_approval_id
                    )
                    approval_hierarchy["level3"].update(
                        {
                            "status": "Approved",
                            "action_by": level3_approver_action.username,
                            "timestamp": entry.level_3_approval_timestamp,
                        }
                    )
                except User.DoesNotExist:
                    approval_hierarchy["level3"].update(
                        {
                            "action_by": "Unknown User",
                            "timestamp": entry.level_3_approval_timestamp,
                        }
                    )

        # Rejection info
        rejection_info = None
        if entry.rejected_by_id:
            try:
                rejector = User.objects.get(id=entry.rejected_by_id)
                rejection_info = {
                    "rejected_by": rejector.username,
                    "reason": entry.rejection_reason,
                    "timestamp": entry.rejection_timestamp,
                }

                # Update the status of the relevant level to "Rejected"
                if entry.level_1_approval_id and not entry.level_2_approval_id:
                    # Rejected by Level 2
                    if "level2" in approval_hierarchy:
                        approval_hierarchy["level2"].update(
                            {
                                "status": "Rejected",
                                "action_by": rejector.username,
                                "timestamp": entry.rejection_timestamp,
                            }
                        )
                elif (
                    entry.level_2_approval_id
                    and not entry.level_3_approval_id
                    and entry.requires_level_3_approval
                ):
                    # Rejected by Level 3
                    if "level3" in approval_hierarchy:
                        approval_hierarchy["level3"].update(
                            {
                                "status": "Rejected",
                                "action_by": rejector.username,
                                "timestamp": entry.rejection_timestamp,
                            }
                        )
                else:
                    # Rejected by Level 1
                    if "level1" in approval_hierarchy:
                        approval_hierarchy["level1"].update(
                            {
                                "status": "Rejected",
                                "action_by": rejector.username,
                                "timestamp": entry.rejection_timestamp,
                            }
                        )
            except User.DoesNotExist:
                rejection_info = {
                    "rejected_by": "Unknown User",
                    "reason": entry.rejection_reason,
                    "timestamp": entry.rejection_timestamp,
                }

        return render(
            request,
            "checksheet/view_filled_startersheet.html",
            {
                "filled_entries": filled_entries,
                "sheet_name": sheet_name,
                "parameter_thresholds": parameter_thresholds,
                "approval_hierarchy": approval_hierarchy,
                "rejection_info": rejection_info,
                "approval_status": entry.approval_status,
                "sheet_id": entry.id,
            },
        )
    return render(request, "checksheet/access_denied.html")


# ----------------------------------------- To aprove a startersheet--------------------------------#
@login_required
def approve_startersheet(request):
    if request.method != "POST":
        return redirect("acknowledgment_list")

    sheet_id = request.POST.get("sheet_id")
    action = request.POST.get("action")  # 'approve' or 'reject'
    level = int(request.POST.get("level"))  # 1, 2, or 3

    try:
        sheet = FilledStarterSheet.objects.get(id=sheet_id)

        # Verify that the user has permission to approve/reject this sheet
        if not sheet.can_approve(request.user):
            messages.error(request, "You don't have permission to perform this action.")
            return redirect("acknowledgment_list")

        # Handle approval
        if action == "approve":
            if level == 1:
                sheet.level_1_approval = request.user
                sheet.level_1_approval_timestamp = timezone.now()
                sheet.approval_status = "level_1_approved"
                messages.success(request, "Sheet approved at Level 1.")
            elif level == 2:
                sheet.level_2_approval = request.user
                sheet.level_2_approval_timestamp = timezone.now()
                sheet.approval_status = "level_2_approved"

                # If level 3 is not required, mark as completed
                if not sheet.requires_level_3_approval:
                    sheet.approval_status = "completed"
                    messages.success(request, "Sheet fully approved.")
                else:
                    messages.success(request, "Sheet approved at Level 2.")
            elif level == 3:
                sheet.level_3_approval = request.user
                sheet.level_3_approval_timestamp = timezone.now()
                sheet.approval_status = "completed"
                messages.success(request, "Sheet fully approved.")

        # Handle rejection
        elif action == "reject":
            # Get rejection reason (you'd need to add this to your form)
            rejection_reason = request.POST.get("rejection_reason", "")

            # Update sheet
            sheet.rejected_by = request.user
            sheet.rejection_timestamp = timezone.now()
            sheet.rejection_reason = rejection_reason
            sheet.approval_status = "rejected"

            messages.warning(request, "Sheet has been rejected.")

        # Save changes
        sheet.save()

    except FilledStarterSheet.DoesNotExist:
        messages.error(request, "Sheet not found.")
    except Exception as e:
        messages.error(request, f"An error occurred: {str(e)}")

    return redirect("acknowledgment_list")


# ----------------------------------------- forget password --------------------------------#


def reset_password(request):
    if request.method == "POST":
        new_password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if new_password == confirm_password:
            email = request.session.get("reset_email")
            user = User.objects.filter(email=email).first()

            if user:
                user.set_password(new_password)
                user.save()
                messages.success(request, "Password reset successful! Please login.")
                return redirect("login")
            else:
                messages.error(request, "User not found!")
        else:
            messages.error(request, "Passwords do not match!")

    return render(request, "checksheet/reset_password.html")


# --------------- Assign sheet function to asign sheet in userlist by action button ---------------------#


@login_required
def assign_sheets(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)
    checksheets = CheckSheet.objects.all()
    Starter = StarterSheet.objects.all()

    if request.method == "POST":
        selected_check_sheets = request.POST.getlist("check_sheets")
        selected_starter_sheets = request.POST.getlist("starter_sheets")

        # Assign selected sheets and POC PDFs to the user
        user.assigned_check_sheets.set(selected_check_sheets)
        user.assigned_starter_sheets.set(selected_starter_sheets)

        messages.success(request, f"Assignments updated for {user.username}")
        return redirect("user_list")

    return render(
        request,
        "checksheet/assign_sheets.html",
        {
            "user": user,
            "checksheets": checksheets,
            "Starter": Starter,
        },
    )


def parse_date(date_str):
    """Helper function to parse date from string."""
    if not date_str:  # Handle None case
        return None
    try:
        return datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return None


# ----------------------------------------- reporting function  --------------------------------#


@login_required
def report_view(request):
    if request.user.role == "admin" or has_page_access(request.user, "report"):
        # Get all checksheets and starter sheets
        checksheets = CheckSheet.objects.all()
        starters = StarterSheet.objects.all()

        # Get distinct lines from checksheets
        distinct_lines = CheckSheet.objects.values_list("line", flat=True).distinct()

        # Get selected line from request
        selected_line = request.GET.get("line")

        # Filter checksheets by line if specified
        if selected_line:
            checksheets = checksheets.filter(line=selected_line)
            starters = starters.filter(
                line=selected_line
            )  # assuming StarterSheet also has a line field

        selected_checksheet = request.GET.get("checksheet")
        selected_startersheet = request.GET.get("startersheet")

        # Safely parse start_date and end_date
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")

        try:
            start_date = parse_date(start_date_str) if start_date_str else None
        except (ParserError, ValueError):
            start_date = None

        try:
            end_date = parse_date(end_date_str) if end_date_str else None
        except (ParserError, ValueError):
            end_date = None

        shift = request.GET.get("shift")
        tab = request.GET.get("tab", "checksheet")
        print(tab)
        report_data = []
        rejection_reasons = []

        # Only proceed if checksheet or startersheet is selected
        if selected_checksheet or selected_startersheet:
            get_image = (
                CheckSheetImage.objects.filter(checksheet=selected_checksheet)
                if selected_checksheet
                else []
            )
            from datetime import time as datetime_time

            # Only proceed with filtering if both dates are provided
            if start_date and end_date:
                end_date = datetime.combine(
                    end_date, datetime_time(hour=23, minute=59, second=59)
                )
                # Rest of your filtering logic remains the same
                if tab == "checksheet" and selected_checksheet:
                    entries = FilledCheckSheet.objects.filter(
                        checksheet_id=selected_checksheet,
                        timestamp__range=(start_date, end_date),
                    )
                    if shift:
                        entries = entries.filter(shift=shift)

                    consolidated_data = {}
                    rejection_details = defaultdict(list)
                    approval_info = {}  # Store approval information for each key

                    for entry in entries:
                        key = (entry.user.username, entry.timestamp.date(), entry.shift)
                        if key not in consolidated_data:
                            consolidated_data[key] = defaultdict(int)

                            # Store approval information
                            approval_info[key] = {
                                "approval_status": entry.approval_status,
                                "level_1_approval_id": entry.level_1_approval_id,
                                "level_1_approval_timestamp": entry.level_1_approval_timestamp,
                                "level_2_approval_id": entry.level_2_approval_id,
                                "level_2_approval_timestamp": entry.level_2_approval_timestamp,
                                "level_3_approval_id": entry.level_3_approval_id,
                                "level_3_approval_timestamp": entry.level_3_approval_timestamp,
                                "rejected_by_id": entry.rejected_by_id,
                                "rejection_timestamp": entry.rejection_timestamp,
                            }

                        for k, v in entry.status_data.items():
                            if v == "Yes":
                                consolidated_data[key][k] += 1
                            elif isinstance(v, int):
                                consolidated_data[key][k] += v
                            elif k == "completely_reject":
                                if isinstance(v, int):
                                    consolidated_data[key][k] += v
                                elif isinstance(v, str) and v.strip():
                                    consolidated_data[key][k] += 1
                                    rejection_details[key].append(v)

                    for (user, date, shift), status_counts in consolidated_data.items():
                        status_display = ", ".join(
                            f"{k}: {v}" for k, v in status_counts.items()
                        )

                        # Get approval info for this entry
                        info = approval_info.get((user, date, shift), {})

                        # Determine who acknowledged and when
                        acknowledged_by = "N/A"

                        # Check if rejected
                        if info.get("rejected_by_id"):
                            # You'll need to fetch the username for the rejected_by_id
                            rejected_by_user = User.objects.filter(
                                id=info.get("rejected_by_id")
                            ).first()
                            if rejected_by_user:
                                acknowledged_by = (
                                    f"Rejected by {rejected_by_user.username}"
                                )
                            else:
                                acknowledged_by = "Rejected"

                        # Check highest level of approval
                        elif info.get("level_3_approval_id"):
                            level3_user = User.objects.filter(
                                id=info.get("level_3_approval_id")
                            ).first()
                            if level3_user:
                                acknowledged_by = f"Level 3: {level3_user.username}"

                        elif info.get("level_2_approval_id"):
                            level2_user = User.objects.filter(
                                id=info.get("level_2_approval_id")
                            ).first()
                            if level2_user:
                                acknowledged_by = f"Level 2: {level2_user.username}"

                        elif info.get("level_1_approval_id"):
                            level1_user = User.objects.filter(
                                id=info.get("level_1_approval_id")
                            ).first()
                            if level1_user:
                                acknowledged_by = f"Level 1: {level1_user.username}"

                        report_data.append(
                            {
                                "user": user,
                                "timestamp": date.strftime("%Y-%m-%d"),
                                "shift": shift,
                                "acknowledgment": info.get("approval_status", "N/A"),
                                "acknowledged_by": acknowledged_by,
                                "status_data": status_display,
                            }
                        )
                        if (user, date, shift) in rejection_details:
                            rejection_reasons.append(
                                {
                                    "user": user,
                                    "timestamp": date.strftime("%Y-%m-%d"),
                                    "shift": shift,
                                    "reasons": ", ".join(
                                        rejection_details[(user, date, shift)]
                                    ),
                                }
                            )

                elif tab == "startersheet" and selected_startersheet:
                    print("in startsheeet")
                    entries = FilledStarterSheet.objects.filter(
                        startersheet_id=selected_startersheet,
                        timestamp__range=(start_date, end_date),
                    )
                    print(entries, "effffffffffffffffff")
                    if shift:
                        entries = entries.filter(shift=shift)

                    for entry in entries:
                        # Determine the acknowledgment status details
                        acknowledged_by = "N/A"
                        acknowledged_time = None

                        # Check if rejected
                        if entry.rejected_by_id:
                            acknowledged_by = (
                                f"Rejected by {entry.rejected_by.username}"
                            )
                            acknowledged_time = entry.rejection_timestamp
                        # Check highest level of approval
                        elif entry.level_3_approval_id:
                            acknowledged_by = (
                                f"Level 3: {entry.level_3_approval.username}"
                            )
                            acknowledged_time = entry.level_3_approval_timestamp
                        elif entry.level_2_approval_id:
                            acknowledged_by = (
                                f"Level 2: {entry.level_2_approval.username}"
                            )
                            acknowledged_time = entry.level_2_approval_timestamp
                        elif entry.level_1_approval_id:
                            acknowledged_by = (
                                f"Level 1: {entry.level_1_approval.username}"
                            )
                            acknowledged_time = entry.level_1_approval_timestamp

                        # Format the timestamp if it exists
                        acknowledged_time_str = (
                            acknowledged_time.strftime("%Y-%m-%d %H:%M:%S")
                            if acknowledged_time
                            else "N/A"
                        )

                        report_data.append(
                            {
                                "user": entry.filled_by.username,
                                "timestamp": entry.timestamp.strftime(
                                    "%Y-%m-%d %H:%M:%S"
                                ),
                                "shift": entry.shift,
                                "acknowledgment": entry.approval_status,
                                "acknowledged_by": acknowledged_by,
                                "acknowledged_time": acknowledged_time_str,
                                "status_data": ", ".join(
                                    f"{k}: {v}" for k, v in entry.status_data.items()
                                ),
                            }
                        )

        if request.GET.get("download") == "pdf":
            buffer = BytesIO()
            pdf = canvas.Canvas(buffer, pagesize=letter)
            width, height = letter

            checksheet_name = "N/A"
            if tab == "checksheet" and selected_checksheet:
                try:
                    checksheet = CheckSheet.objects.get(id=selected_checksheet)
                    checksheet_name = checksheet.name
                except CheckSheet.DoesNotExist:
                    pass
            elif tab == "startersheet" and selected_startersheet:
                try:
                    startersheet = StarterSheet.objects.get(id=selected_startersheet)
                    checksheet_name = startersheet.name
                except StarterSheet.DoesNotExist:
                    pass

            pdf.setTitle("Report")
            report_type = f"{checksheet_name} Report"
            start_date_str = start_date.strftime("%Y-%m-%d") if start_date else "N/A"
            end_date_str = end_date.strftime("%Y-%m-%d") if end_date else "N/A"

            pdf.setFont("Helvetica-Bold", 16)
            title_width = pdf.stringWidth(report_type, "Helvetica-Bold", 16)
            pdf.drawString((width - title_width) / 2, height - 50, report_type)

            pdf.setFont("Helvetica", 12)
            date_text = f"From: {start_date_str}  To: {end_date_str}"
            date_width = pdf.stringWidth(date_text, "Helvetica", 12)
            pdf.drawString((width - date_width) / 2, height - 70, date_text)

            y_position = height - 100

            if tab == "checksheet" and get_image.exists():
                pdf.setFont("Helvetica-Bold", 12)
                pdf.drawString(40, y_position, "Images:")
                y_position -= 20

                image_count = min(len(get_image), 3)
                if image_count > 0:
                    max_img_width = min(150, (width - 80) / image_count)
                    max_img_height = 120
                    total_width = image_count * max_img_width
                    start_x = (width - total_width) / 2

                    for idx, img in enumerate(get_image[:image_count]):
                        try:
                            img_path = img.image.path
                            x_pos = start_x + (idx * max_img_width)
                            pdf.drawImage(
                                img_path,
                                x_pos,
                                y_position - max_img_height,
                                width=max_img_width,
                                height=max_img_height,
                                preserveAspectRatio=True,
                            )
                        except Exception as e:
                            print(f"Error adding image to PDF: {e}")
                    y_position -= max_img_height + 30

            y_position -= 20

            table_width = width - 80
            table_left = 40
            col_widths_percent = [15, 15, 10, 15, 20, 25]
            col_widths = [int(table_width * w / 100) for w in col_widths_percent]
            col_positions = [table_left]
            for w in col_widths[:-1]:
                col_positions.append(col_positions[-1] + w)

            header_height = 20
            pdf.setLineWidth(1)
            pdf.setFillColorRGB(0.9, 0.9, 0.9)
            pdf.rect(table_left, y_position, table_width, -header_height, fill=1)
            pdf.setFillColorRGB(0, 0, 0)

            headers = [
                "User",
                "Timestamp",
                "Shift",
                "Approved",
                "Approved By",
                "Data",
            ]
            pdf.setFont("Helvetica-Bold", 10)
            for i, header in enumerate(headers):
                x = col_positions[i] + 2
                pdf.drawString(x, y_position - 14, header)

            y_position -= header_height
            pdf.setFont("Helvetica", 9)

            for idx, entry in enumerate(report_data):
                status_data = entry["status_data"]
                timestamp_display = entry["timestamp"]
                if tab == "startersheet" and " " in timestamp_display:
                    timestamp_display = timestamp_display.split(" ")[0]

                # Append rejection reasons for checksheet tab
                if (
                    tab == "checksheet"
                    and idx < len(rejection_reasons)
                    and rejection_reasons[idx]["timestamp"] == entry["timestamp"]
                ):
                    status_data += (
                        f" (Rejection Reasons: {rejection_reasons[idx]['reasons']})"
                    )

                data_width = col_widths[-1] - 4
                char_width = pdf.stringWidth("x", "Helvetica", 9)
                chars_per_line = max(1, int(data_width / char_width))
                lines_needed = max(
                    1,
                    len(status_data) // chars_per_line
                    + (1 if len(status_data) % chars_per_line > 0 else 0),
                )
                lines_needed = min(lines_needed, 5)
                row_height = max(20, lines_needed * 12)

                if y_position - row_height < 50:
                    pdf.showPage()
                    pdf.setFont("Helvetica-Bold", 16)
                    pdf.drawString((width - title_width) / 2, height - 50, report_type)
                    y_position = height - 100
                    pdf.setFillColorRGB(0.9, 0.9, 0.9)
                    pdf.rect(
                        table_left, y_position, table_width, -header_height, fill=1
                    )
                    pdf.setFillColorRGB(0, 0, 0)
                    pdf.setFont("Helvetica-Bold", 10)
                    for i, header in enumerate(headers):
                        x = col_positions[i] + 2
                        pdf.drawString(x, y_position - 14, header)
                    y_position -= header_height

                pdf.setLineWidth(0.5)
                pdf.rect(table_left, y_position, table_width, -row_height)
                for x in col_positions[1:]:
                    pdf.line(x, y_position, x, y_position - row_height)

                cell_data = [
                    entry["user"],
                    timestamp_display,
                    entry["shift"],
                    entry["acknowledgment"],
                    entry["acknowledged_by"],
                    status_data,
                ]
                for i in range(5):
                    x = col_positions[i] + 2
                    pdf.drawString(x, y_position - 14, cell_data[i])

                data_x = col_positions[5] + 2
                data_y = y_position - 12
                if len(status_data) > chars_per_line:
                    text_object = pdf.beginText(data_x, data_y)
                    text_object.setFont("Helvetica", 9)
                    words = status_data.split()
                    current_line = ""
                    for word in words:
                        test_line = current_line + " " + word if current_line else word
                        if pdf.stringWidth(test_line, "Helvetica", 9) < data_width:
                            current_line = test_line
                        else:
                            text_object.textLine(current_line)
                            current_line = word
                            if text_object.getY() < data_y - (lines_needed - 1) * 12:
                                current_line += "..."
                                break
                    if current_line:
                        text_object.textLine(current_line)
                    pdf.drawText(text_object)
                else:
                    pdf.drawString(data_x, data_y, status_data)

                y_position -= row_height

            table_height = (height - 100 - y_position) - 20
            pdf.setLineWidth(1.5)
            pdf.rect(table_left, height - 100 - 20, table_width, -table_height)
            pdf.setFont("Helvetica", 8)
            pdf.drawString(width - 60, 30, f"Page {pdf.getPageNumber()}")
            pdf.save()
            buffer.seek(0)
            response = HttpResponse(buffer, content_type="application/pdf")
            filename = f"{checksheet_name} {start_date_str} to {end_date_str}".replace(
                " ", "_"
            ).replace("/", "-")
            response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
            return response

        if request.GET.get("download") == "excel":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            medium_border = Border(
                left=Side(style="medium"),
                right=Side(style="medium"),
                top=Side(style="medium"),
                bottom=Side(style="medium"),
            )
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )
            subheader_fill = PatternFill(
                start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
            )
            title_fill = PatternFill(
                start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
            )
            alt_row_fill = PatternFill(
                start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
            )

            checksheet_name = "N/A"
            if tab == "checksheet" and selected_checksheet:
                try:
                    checksheet = CheckSheet.objects.get(id=selected_checksheet)
                    checksheet_name = checksheet.name
                except CheckSheet.DoesNotExist:
                    pass
            elif tab == "startersheet" and selected_startersheet:
                try:
                    startersheet = StarterSheet.objects.get(id=selected_startersheet)
                    checksheet_name = startersheet.name
                except StarterSheet.DoesNotExist:
                    pass

            ws.cell(row=1, column=1).value = f"Report: {checksheet_name}"
            ws.cell(row=1, column=1).font = Font(bold=True, size=16, color="000000")
            ws.merge_cells("A1:F1")
            ws.cell(row=1, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=1, column=1).border = medium_border
            ws.cell(row=1, column=1).fill = title_fill
            ws.row_dimensions[1].height = 30

            date_range = f"Period: {start_date.strftime('%Y-%m-%d') if start_date else 'N/A'} to {end_date.strftime('%Y-%m-%d') if end_date else 'N/A'}"
            ws.cell(row=2, column=1).value = date_range
            ws.cell(row=2, column=1).font = Font(bold=True, italic=True)
            ws.merge_cells("A2:F2")
            ws.cell(row=2, column=1).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=2, column=1).border = thin_border
            ws.cell(row=2, column=1).fill = subheader_fill
            ws.row_dimensions[2].height = 20

            current_row = 4

            if tab == "checksheet" and get_image.exists():
                ws.cell(row=current_row, column=1).value = "Images:"
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                ws.cell(row=current_row, column=1).border = thin_border
                ws.cell(row=current_row, column=1).fill = subheader_fill
                for col in range(2, 7):
                    ws.cell(row=current_row, column=col).border = thin_border
                current_row += 1
                img_row = current_row
                ws.row_dimensions[img_row].height = (
                    120  # Increased to ensure images fit
                )
                image_count = min(len(get_image), 4)
                start_col = (
                    2  # Start from column B to avoid overlapping with "Images:" label
                )

                for img_idx in range(image_count):

                    try:
                        # Custom column placement
                        if img_idx == 0:
                            col_num = 1  # First image in column A
                        elif img_idx == 1:
                            col_num = 3  # Second image in column C
                        elif img_idx == 2:
                            col_num = 5  # Third image in column E
                        else:
                            col_num = 6  # Any additional images in column F

                        img = get_image[img_idx]
                        img_path = img.image.path
                        pil_img = PILImage.open(img_path)
                        max_size = (200, 150)  # Keep your smaller size
                        pil_img.thumbnail(max_size)
                        img_io = BytesIO()
                        pil_img.save(img_io, format=pil_img.format)
                        img_io.seek(0)
                        xl_img = XLImage(img_io)

                        # Calculate column letter
                        col_letter = chr(64 + col_num)
                        cell_address = f"{col_letter}{img_row}"

                        # Add image to worksheet
                        ws.add_image(xl_img, cell_address)

                        # Make the column wider to avoid overlap
                        ws.column_dimensions[col_letter].width = 30

                        # If you want empty columns between images to look nicer
                        if img_idx < image_count - 1:
                            # Set width for spacing column (column between images)
                            spacing_col = chr(65 + col_num)  # Next column
                            ws.column_dimensions[spacing_col].width = 5

                        print(
                            f"Placed image {img_idx} at column {col_letter} ({cell_address})"
                        )
                    except Exception as e:
                        print(f"Error adding image to Excel: {e}")
                        ws.cell(row=img_row, column=col_num).value = (
                            "Error loading image"
                        )
                        ws.cell(row=img_row, column=col_num).border = thin_border
                current_row = img_row + 2
            else:
                current_row += 1

            headers = [
                "User",
                "Timestamp",
                "Shift",
                "Approved",
                "Approved By",
                "Data",
            ]
            ws.row_dimensions[current_row].height = 28
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.border = medium_border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                if header == "User":
                    ws.column_dimensions[chr(64 + col_num)].width = 15
                elif header == "Timestamp":
                    ws.column_dimensions[chr(64 + col_num)].width = 20
                elif header == "Shift":
                    ws.column_dimensions[chr(64 + col_num)].width = 10
                elif header == "Approved":
                    ws.column_dimensions[chr(64 + col_num)].width = 16
                elif header == "Approved By":
                    ws.column_dimensions[chr(64 + col_num)].width = 18
                elif header == "Data":
                    ws.column_dimensions[chr(64 + col_num)].width = 40

            # Calculate maximum length for Data column
            max_data_length = 40
            for row_num, entry in enumerate(report_data, 1):
                status_data = entry["status_data"]
                if (
                    tab == "checksheet"
                    and row_num - 1 < len(rejection_reasons)
                    and rejection_reasons[row_num - 1]["timestamp"]
                    == entry["timestamp"]
                ):
                    status_data += f" (Rejection Reasons: {rejection_reasons[row_num - 1]['reasons']})"
                max_data_length = max(max_data_length, len(status_data))

            data_column = chr(64 + 6)
            ws.column_dimensions[data_column].width = min(max_data_length * 0.8, 100)

            for row_num, entry in enumerate(report_data, 1):
                data_row = current_row + row_num
                status_data = entry["status_data"]
                if (
                    tab == "checksheet"
                    and row_num - 1 < len(rejection_reasons)
                    and rejection_reasons[row_num - 1]["timestamp"]
                    == entry["timestamp"]
                ):
                    status_data += f" (Rejection Reasons: {rejection_reasons[row_num - 1]['reasons']})"

                data_width = ws.column_dimensions[data_column].width * 1.2
                lines_needed = max(
                    1,
                    len(status_data) // int(data_width)
                    + (1 if len(status_data) % int(data_width) > 0 else 0),
                )
                row_height = max(22, lines_needed * 15)
                ws.row_dimensions[data_row].height = row_height

                row_fill = alt_row_fill if row_num % 2 == 0 else None

                data_cells = [
                    entry["user"],
                    entry["timestamp"],
                    entry["shift"],
                    entry["acknowledgment"],
                    entry["acknowledged_by"],
                    status_data,
                ]

                for col_num, value in enumerate(data_cells, 1):
                    cell = ws.cell(row=data_row, column=col_num)
                    cell.value = value
                    cell.border = thin_border
                    if row_fill:
                        cell.fill = row_fill
                    if col_num in [1, 3, 4, 5]:
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                    elif col_num == 2:
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        if isinstance(value, str) and len(value) > 10:
                            try:
                                cell.number_format = "yyyy-mm-dd hh:mm"
                            except:
                                pass
                    else:
                        cell.alignment = Alignment(vertical="center", wrap_text=True)

            footer_row = current_row + len(report_data) + 2
            ws.cell(row=footer_row, column=1).value = (
                f"Report generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )
            ws.cell(row=footer_row, column=1).font = Font(italic=True, size=9)
            ws.merge_cells(f"A{footer_row}:F{footer_row}")
            ws.cell(row=footer_row, column=1).alignment = Alignment(horizontal="right")

            excel_io = BytesIO()
            wb.save(excel_io)
            excel_io.seek(0)
            response = HttpResponse(
                excel_io.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            start_date_str = start_date.strftime("%d-%m-%Y") if start_date else "N/A"
            end_date_str = end_date.strftime("%d-%m-%Y") if end_date else "N/A"
            filename = f"{checksheet_name} {start_date_str} to {end_date_str}".replace(
                " ", "_"
            ).replace("/", "-")
            response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
            return response

        print(report_data, "repor")
        return render(
            request,
            "checksheet/report.html",
            {
                "checksheets": checksheets,
                "starters": starters,
                "report_data": report_data,
                "tab": tab,
                "distinct_lines": distinct_lines,
            },
        )

    return render(request, "checksheet/access_denied.html")


# ----------------------------------------- in fill_checksheet edit button to get data --------------------------------#
def get_today_checksheets(request):
    today = localdate()
    checksheet_id = request.GET.get("checksheet_id")
    line = request.GET.get("line")
    shift = request.GET.get("shift")

    print(checksheet_id, line)
    checksheets = FilledCheckSheet.objects.filter(
        checksheet=checksheet_id, line=line, timestamp__date=today, shift=shift
    )

    data = [
        {"id": c.id, "status_data": c.status_data, "timestamp": c.timestamp}
        for c in checksheets
    ]
    print(data)

    return JsonResponse(data, safe=False)


# ----------------------------------------- function to request password change by user   --------------------------------#
def request_password_reset(request):
    if request.method == "POST":
        employee_id = request.POST.get("employee_id")
        user = CustomUser.objects.filter(employee_id=employee_id).first()

        if user:
            # Check if a request already exists
            existing_request = PasswordResetRequest.objects.filter(
                user=user, status="pending"
            ).first()
            if existing_request:
                messages.info(request, "A reset request is already pending.")
            else:
                # Create a new request
                PasswordResetRequest.objects.create(user=user)
                messages.success(request, "Password reset request sent to admin.")

            return redirect("login")  # Redirect to login page
        else:
            messages.error(request, "Invalid Employee ID.")

    return render(request, "checksheet/request_password_reset.html")


# ----------------------------------------- show passwod request to admin  --------------------------------#
@user_passes_test(is_admin)
def admin_password_requests(request):
    requests = PasswordResetRequest.objects.filter(status="pending")
    pending_requests_count = requests.count()  # Count pending requests
    checksheets = CheckSheet.objects.all()
    Starter = StarterSheet.objects.all()

    return render(
        request,
        "checksheet/admin_password_requests.html",
        {
            "requests": requests,
            "pending_requests_count": pending_requests_count,
            "checksheets": checksheets,
            "Starter": Starter,
        },
    )


# ----------------------------------------- aprove password request by admin  --------------------------------#


@user_passes_test(is_admin)
def approve_password_reset(request, request_id):
    reset_request = PasswordResetRequest.objects.get(id=request_id)

    if request.method == "POST":
        new_password = request.POST.get("new_password")

        if new_password:
            reset_request.user.set_password(new_password)  # Set new password
            reset_request.user.save()
            reset_request.status = "approved"
            reset_request.save()
            messages.success(request, "Password reset successfully!")

            return redirect("admin_password_requests")

    return render(
        request,
        "checksheet/approve_password_reset.html",
        {"reset_request": reset_request},
    )


# ----------------------------------------- manage acces function to give feature to users  --------------------------------#
@login_required
@user_passes_test(lambda user: user.role == "admin")  # Only admin can access
def manage_access(request):
    users = User.objects.exclude(role="admin")  # Exclude admin itself

    pages = {
        "all_checksheets": "CheckSheet",
        "upload_poc": "OPS",
        "report": "Report",
        "all_startersheet": "StarterSheet",
        "acknowledgment_list": "Acknowledgment",
        "fill_starter_sheet": "Fill Starter Sheet",
        "fill_checksheet_detail": "Fill CheckSheet",
        "form_request_view": "Request Back Data",
        "home": "Dashboard",
    }

    # Fetch all access records
    access_data = PageAccess.objects.filter(user__in=users)

    # Create a dictionary mapping user to their allowed pages
    user_access = {user.id: set() for user in users}
    for access in access_data:
        if access.has_access:
            user_access[access.user.id].add(access.page_name)

    # Attach access pages to users
    for user in users:
        user.access_pages = user_access.get(user.id, set())

    if request.method == "POST":
        for user in users:
            for page in pages.keys():
                field_name = f"access_{user.id}_{page}"
                has_access = request.POST.get(field_name) == "1"

                try:
                    # Try to get a single object
                    obj = PageAccess.objects.get(user=user, page_name=page)
                    obj.has_access = has_access
                    obj.save()
                except PageAccess.DoesNotExist:
                    # Create if it doesn't exist
                    PageAccess.objects.create(
                        user=user, page_name=page, has_access=has_access
                    )
                except PageAccess.MultipleObjectsReturned:
                    # Handle multiple objects scenario
                    # First, update all existing records to the same value
                    PageAccess.objects.filter(user=user, page_name=page).update(
                        has_access=has_access
                    )

                    # Then, if you want to clean up duplicates, keep only one record
                    duplicates = PageAccess.objects.filter(
                        user=user, page_name=page
                    ).order_by("id")
                    first_record = duplicates.first()
                    # Delete all other records except the first one
                    duplicates.exclude(id=first_record.id).delete()
        messages.success(request, "Access Granted", extra_tags="manage")
        return redirect(manage_access)

    return render(
        request,
        "checksheet/manage_access.html",
        {
            "users": users,
            "pages": pages,
        },
    )


# ----------------------------------------- function to delect filled checksheet data by operator  --------------------------------#
@csrf_exempt
def delete_checksheet(request, checksheet_id):
    if request.method == "DELETE":
        try:
            checksheet = FilledCheckSheet.objects.get(id=checksheet_id)
            checksheet.delete()
            return JsonResponse({"success": True})
        except FilledCheckSheet.DoesNotExist:
            return JsonResponse({"success": False, "error": "Checksheet not found"})
    return JsonResponse({"success": False, "error": "Invalid request method"})


# ----------------------------------------- show checksheet data pending acknowledgement  --------------------------------#
@login_required
def pending_acknowledgments_check(request):
    pending_requests = FormRequest.objects.filter(status="Pending")
    pending_acknowledgments = (
        FilledStarterSheet.objects.filter(acknowledgment="No")
        .values(
            "startersheet__id",
            "startersheet__name",
            "filled_by__id",
            "filled_by__username",
            "shift",
        )
        .distinct()
    )

    shift = request.GET.get("shift", "")
    selected_date = request.GET.get("selected_date", "")

    if not shift and not selected_date:
        return render(
            request,
            "checksheet/acknowledge.html",
            {"pending_acknowledgments_checksheet": []},
        )

    pending_acknowledgments_checksheet = FilledCheckSheet.objects.filter(
        acknowledgment="No"
    )

    if shift:
        pending_acknowledgments_checksheet = pending_acknowledgments_checksheet.filter(
            shift=shift
        )

    if selected_date:
        selected_date = parse_date(selected_date)
        if selected_date:
            pending_acknowledgments_checksheet = (
                pending_acknowledgments_checksheet.filter(timestamp__date=selected_date)
            )

    # Consolidated Data
    consolidated_data = {}

    for entry in pending_acknowledgments_checksheet:
        key = (entry.checksheet.name, entry.user.username, entry.shift)
        if key not in consolidated_data:
            consolidated_data[key] = defaultdict(int)
            consolidated_data[key]["reject_reasons"] = set()

        for k, v in entry.status_data.items():
            if v == "Yes":
                consolidated_data[key][k] += 1
            elif isinstance(v, int):
                consolidated_data[key][k] += v
            elif k == "completely_reject":
                if isinstance(v, int):  # If it's numeric, sum it
                    consolidated_data[key][k] += v
                else:
                    consolidated_data[key][
                        "completely_reject"
                    ] += 1  # Convert text rejection to count

    # Convert to structured list
    consolidated_entries = []
    for (checksheet_name, username, shift), status_counts in consolidated_data.items():
        status_display = ", ".join(
            f"{k}: {v}" for k, v in status_counts.items() if k != "reject_reasons"
        )
        reject_reasons = (
            ", ".join(status_counts["reject_reasons"])
            if status_counts["reject_reasons"]
            else ""
        )

        if reject_reasons:
            status_display += f", completely_reject: {reject_reasons}"

        consolidated_entries.append(
            {
                "checksheet_name": checksheet_name,
                "username": username,
                "shift": shift,
                "acknowledgment": "Pending",
                "status_data": status_display,
                "pending_acknowledgments": pending_acknowledgments,
            }
        )

    return render(
        request,
        "checksheet/acknowledge.html",
        {
            "pending_acknowledgments_checksheet": consolidated_entries,
            "pending_acknowledgments": pending_acknowledgments,
            "pending_requests": pending_requests,
        },
    )


# ----------------------------------------- function for acknowledge checksheet  accept or reject  --------------------------------#


@login_required
def acknowledge_checksheets(request):
    if request.method != "POST":
        return redirect("acknowledgment_list")

    action = request.POST.get("action")  # 'accept' or 'reject'
    selected_checksheets = request.POST.getlist("selected_checksheets")
    shift = request.POST.get("shift")
    selected_date = request.POST.get("selected_date")
    active_tab = request.POST.get("active_tab", "checksheet")
    print(action, selected_checksheets, shift, selected_date, active_tab)

    if not selected_checksheets:
        messages.warning(request, "No checksheets selected.")
        return redirect("acknowledgment_list")

    processed_count = 0

    for checksheet_data in selected_checksheets:
        try:
            # Format is checksheet_name|username|shift|date
            checksheet_name, username, sheet_shift, date_str = checksheet_data.split(
                "|"
            )

            # Convert string date to datetime object
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")

            # Get all matching checksheets for this user, date, shift and name
            # Correctly use the related fields based on your model structure
            checksheets = FilledCheckSheet.objects.filter(
                checksheet__name=checksheet_name,  # Use __ to access related field
                user__username=username,  # Use __ to access related field
                shift=sheet_shift,
                timestamp__date=date_obj.date(),  # Match the date part of timestamp
            )

            # Check if any checksheets exist
            if not checksheets.exists():
                messages.error(
                    request, f"No checksheets found for {checksheet_name} on {date_str}"
                )
                continue

            # Process approval/rejection
            for checksheet in checksheets:
                # Handle approval permission check and update directly
                if action == "accept":
                    # First level approval
                    if (
                        checksheet.approval_status == "pending"
                        and checksheet.assigned_level_1_approver  # Make sure it's not None
                        and request.user.id == checksheet.assigned_level_1_approver.id
                    ):
                        checksheet.level_1_approval = (
                            request.user
                        )  # Assign the user object, not just ID
                        checksheet.level_1_approval_timestamp = timezone.now()
                        checksheet.approval_status = "level_1_approved"
                        checksheet.save()
                        processed_count += 1

                    # Second level approval
                    elif (
                        checksheet.approval_status == "level_1_approved"
                        and checksheet.assigned_level_2_approver  # Make sure it's not None
                        and request.user.id == checksheet.assigned_level_2_approver.id
                    ):
                        checksheet.level_2_approval = (
                            request.user
                        )  # Assign the user object, not just ID
                        checksheet.level_2_approval_timestamp = timezone.now()
                        checksheet.approval_status = "level_2_approved"

                        # Check if level 3 is required
                        if not checksheet.requires_level_3_approval:
                            checksheet.approval_status = "completed"

                        checksheet.save()
                        processed_count += 1

                    # Third level approval (admin only)
                    elif (
                        checksheet.approval_status == "level_2_approved"
                        and request.user.role == "admin"  # Make sure role field exists
                    ):
                        checksheet.level_3_approval = (
                            request.user
                        )  # Assign the user object, not just ID
                        checksheet.level_3_approval_timestamp = timezone.now()
                        checksheet.approval_status = "completed"
                        checksheet.save()
                        processed_count += 1
                # Handle rejection
                elif action == "reject":
                    # Check if user has permission to reject
                    can_reject = False

                    if (
                        checksheet.approval_status == "pending"
                        and hasattr(checksheet, "assigned_level_1_approver")
                        and request.user.id == checksheet.assigned_level_1_approver.id
                    ):
                        can_reject = True
                    elif (
                        checksheet.approval_status == "level_1_approved"
                        and hasattr(checksheet, "assigned_level_2_approver")
                        and request.user.id == checksheet.assigned_level_2_approver.id
                    ):
                        can_reject = True
                    elif (
                        checksheet.approval_status == "level_2_approved"
                        and request.user.role == "admin"
                    ):
                        can_reject = True

                    if can_reject:
                        rejection_reason = request.POST.get("rejection_reason", "")

                        checksheet.rejected_by = request.user
                        checksheet.rejection_timestamp = timezone.now()
                        checksheet.rejection_reason = rejection_reason
                        checksheet.approval_status = "rejected"

                        checksheet.save()
                        processed_count += 1

        except ValueError:
            messages.error(
                request, f"Invalid checksheet data format: {checksheet_data}"
            )
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")

    if processed_count > 0:
        if action == "accept":
            messages.success(
                request, f"{processed_count} checksheet(s) accepted successfully."
            )
        else:
            messages.warning(request, f"{processed_count} checksheet(s) rejected.")
    else:
        messages.warning(
            request, "No checksheets were processed. You may not have permission."
        )

    # Redirect back to the list with the same filters
    return redirect("acknowledgment_list")


# ----------------------------------------- function for assign opc in chcksheet page  --------------------------------#


@login_required
def assign_poc_bulk(request):
    startersheets = StarterSheet.objects.prefetch_related("assigned_pocs").all()
    all_pocs = POCUpload.objects.all()

    if request.method == "POST":
        # Clear previous assignments
        for startersheet in startersheets:
            startersheet.assigned_pocs.clear()

        # Assign selected PDFs to CheckSheets
        for startersheet in startersheets:
            selected_poc_ids = request.POST.getlist(
                f"startersheet_{startersheet.id}_poc"
            )
            selected_pocs = POCUpload.objects.filter(id__in=selected_poc_ids)
            startersheet.assigned_pocs.set(selected_pocs)  # Assign PDFs to CheckSheet

        return redirect("all_startersheet")  # Redirect back after assignment

    return render(
        request,
        "checksheet/assign_poc.html",
        {"startersheets": startersheets, "all_pocs": all_pocs},
    )


# ----------------------------------------- back date data request functionality  --------------------------------#


@login_required
def form_request_view(request):
    if request.user.role == "shift_incharge" or has_page_access(
        request.user, "form_request_view"
    ):
        if request.method == "POST":
            number_of_requests = int(request.POST.get("number_of_requests", 1))
            if number_of_requests > 20:
                messages.error(request, "Maximum number of requests is 20.")
                return redirect("form_request_view")

            shifts = request.POST.getlist("shift")
            users = request.POST.getlist("user")
            checksheet_ids = request.POST.getlist("checksheet_ids")
            dates = request.POST.getlist("date")
            visible_untils = request.POST.getlist("visible_until")
            reasons = request.POST.getlist("reasons")
            lines = request.POST.getlist("line")  # New: Get list of selected lines

            if (
                len(shifts) != number_of_requests
                or len(users) != number_of_requests
                or len(checksheet_ids) != number_of_requests
                or len(dates) != number_of_requests
                or len(reasons) != number_of_requests
                or len(lines) != number_of_requests  # New: Validate lines
            ):
                messages.error(
                    request, "Invalid form submission. Please fill all fields."
                )
                return redirect("form_request_view")

            for i in range(number_of_requests):
                visible_until = None
                if visible_untils[i]:
                    dt = datetime.fromisoformat(visible_untils[i])
                    visible_until = make_aware(dt)

                checksheet = CheckSheet.objects.get(id=checksheet_ids[i])
                user = CustomUser.objects.get(id=users[i])
                FormRequest.objects.create(
                    checksheet=checksheet,
                    shift=shifts[i],
                    user=user,
                    date=dates[i],
                    reason=reasons[i],
                    visible_until=visible_until,
                    line=lines[i],  # New: Save the selected line
                )

            messages.success(
                request,
                "Back Data Submitted successfully!",
                extra_tags="back_creation",
            )

        # GET request handling
        checksheets = CheckSheet.objects.all()
        Starter = StarterSheet.objects.all()
        users = [
            {"id": user.id, "username": user.username}
            for user in CustomUser.objects.filter(is_superuser=False)
        ]
        shifts = [("A", "Shift A"), ("B", "Shift B"), ("C", "Shift C")]
        return render(
            request,
            "checksheet/form_request.html",
            {
                "checksheets": checksheets,
                "shifts": shifts,
                "users": users,
                "checksheets_for_js": json.dumps(
                    [
                        {
                            "id": checksheet.id,
                            "name": checksheet.name,
                            "line": checksheet.line,
                        }
                        for checksheet in checksheets
                    ]
                ),
                "shifts_for_js": json.dumps(shifts),
                "users_for_js": json.dumps(users),
                "Starter": Starter,
            },
        )
    return render(request, "checksheet/access_denied.html")


# ----------------------------------------- admin request accept reject for back date data --------------------------------#
@login_required
def accept_request(request, request_id):
    form_request = get_object_or_404(FormRequest, id=request_id)
    form_request.status = "Accepted"
    form_request.save()

    return redirect("acknowledgment_list")


@login_required
def reject_request(request, request_id):
    form_request = get_object_or_404(FormRequest, id=request_id)
    form_request.status = "Rejected"
    form_request.save()
    return redirect("acknowledgment_list")


# ----------------------------------------- back date data filled  by operator --------------------------------#
@login_required
def fill_checksheet_request(request, request_id):
    if request.user.role == "operator":
        form_requests = FormRequest.objects.filter(
            status="Accepted",
            checksheet__assigned_users=request.user,
            visible_until__gte=now(),
        ).select_related("checksheet")

        try:
            form_request = form_requests.get(id=request_id)
        except FormRequest.DoesNotExist:
            return render(request, "checksheet/access_denied.html")

        if request.method == "POST":
            status_data = {}
            complete_reject = request.POST.get("completely_reject")
            zones = Zone.objects.filter(checksheet=form_request.checksheet)
            for zone in zones:
                key = f"zone_{zone.id}"
                value = request.POST.get(key)
                try:
                    value = int(value) if value and value.isdigit() else 0
                except ValueError:
                    value = 0
                status_data[zone.name] = value
            try:
                complete_reject = (
                    int(complete_reject)
                    if complete_reject and complete_reject.isdigit()
                    else 0
                )
            except ValueError:
                complete_reject = 0
            status_data["completely_reject"] = complete_reject

            if any(status_data.values()):  # Check if at least one value is not zero
                try:
                    # Use transaction to ensure data integrity
                    with transaction.atomic():
                        # Find all entries for this checksheet, shift, and date
                        date_obj = (
                            form_request.date
                        )  # Get the date part of the timestamp
                        existing_entries = FilledCheckSheet.objects.filter(
                            checksheet_id=form_request.checksheet.id,
                            shift=form_request.shift,
                            timestamp__date=date_obj,
                            line=form_request.line,
                        )

                        # Delete existing entries if any
                        if existing_entries.exists():
                            existing_entries.delete()

                        # Create new entry
                        FilledCheckSheet.objects.create(
                            checksheet=form_request.checksheet,
                            user=request.user,
                            shift=form_request.shift,
                            timestamp=form_request.date,
                            status_data=status_data,
                            line=form_request.line,
                            send_acknowledgment=True,
                        )

                    # Return JSON response for AJAX
                    return JsonResponse(
                        {"success": True, "message": "Checksheet filled successfully."}
                    )
                except Exception as e:
                    # Return JSON error response for exceptions
                    return JsonResponse({"success": False, "error": str(e)}, status=500)
            else:
                # Return JSON error response
                return JsonResponse(
                    {
                        "success": False,
                        "error": "No valid data provided. Please fill the checksheet properly.",
                    },
                    status=400,
                )

        zones = Zone.objects.filter(checksheet=form_request.checksheet)
        images = form_request.checksheet.images.all()
        return render(
            request,
            "checksheet/fill_checksheet_request.html",
            {
                "form_request": form_request,
                "zones": zones,
                "form_requests": form_requests,
                "images": images,
            },
        )
    return render(request, "checksheet/access_denied.html")


# ----------------------------------------- delete Q-galery(ops/poc) --------------------------------#
def delete_poc(request, poc_id):
    if request.method == "DELETE":
        poc = get_object_or_404(POCUpload, id=poc_id)

        # Remove the PDF file and clear check sheets
        if poc.pdf:
            poc.pdf.delete()  # Delete the PDF file from storage
        poc.assigned_startersheets.clear()  # Remove all assigned check sheets
        poc.delete()  # Delete the POC record from the database

        return JsonResponse({"message": "PDF deleted successfully"}, status=200)

    return JsonResponse({"error": "Invalid request"}, status=400)


# ----------------------------------------- assign checksheet and starter sheet to user by dropdown --------------------------------#
def assign_users(request, sheet_type, sheet_id):
    if request.method == "POST":
        if sheet_type == "checksheet":
            sheet = get_object_or_404(CheckSheet, id=sheet_id)
        elif sheet_type == "startersheet":
            sheet = get_object_or_404(StarterSheet, id=sheet_id)
        else:
            return redirect("all_checksheets")  # Fallback if type is invalid

        user_ids = request.POST.getlist("user_ids")
        users = CustomUser.objects.filter(id__in=user_ids)
        sheet.assigned_users.set(users)

        if sheet_type == "checksheet":
            return redirect("all_checksheets")
        else:
            return redirect("all_startersheet")


@csrf_exempt
@login_required
def mark_poc_as_read(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)  # Load JSON data
            poc_id = data.get("poc_id")  # Get poc_id from JSON
            if poc_id:
                poc = POCUpload.objects.get(id=poc_id)

                # Create a new read status entry for today regardless of previous reads
                today = timezone.now().date()

                # Check if the user has already read this POC today
                already_read_today = POCReadStatus.objects.filter(
                    user=request.user, poc=poc, timestamp__date=today
                ).exists()

                # Only create a new entry if they haven't read it today
                if not already_read_today:
                    POCReadStatus.objects.create(user=request.user, poc=poc, read=True)

                return JsonResponse({"status": "success"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=400)
    return JsonResponse({"status": "error"}, status=400)


# ------------------------------to show images in view checksheet  --------------------------------#


def get_checksheet_images(request):
    checksheet_name = request.GET.get("name", "")
    checksheet_line = request.GET.get("line", "")  # Get the line parameter

    try:
        # Get the checksheet by name AND line
        checksheet = CheckSheet.objects.get(name=checksheet_name, line=checksheet_line)

        # Get all images for this checksheet
        images = CheckSheetImage.objects.filter(checksheet=checksheet)

        # Create a list of image URLs
        image_data = [{"url": image.image.url, "id": image.id} for image in images]

        return JsonResponse({"success": True, "images": image_data})
    except CheckSheet.DoesNotExist:
        return JsonResponse({"success": False, "error": "Checksheet not found"})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})


# ------------------------------when operator click finish to send for acknowledgment  --------------------------------#
@login_required
def send_acknowledgments(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            checksheet_id = data.get("checksheet_id")
            shift = data.get("shift")

            # Get today's date
            today = timezone.now().date()
            print(checksheet_id, shift, today)

            # Update all FilledCheckSheet entries for today, current user, checksheet, and shift
            updated_count = FilledCheckSheet.objects.filter(
                user=request.user,
                checksheet_id=checksheet_id,
                shift=shift,
                timestamp__date=today,
                send_acknowledgment=False,
            ).update(send_acknowledgment=True)

            return JsonResponse(
                {"success": True, "message": f"Updated {updated_count} entries"}
            )
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)}, status=400)
    return JsonResponse(
        {"success": False, "error": "Invalid request method"}, status=405
    )


# ------------------------------ to assing user on approver level on checksheet and startersheet page  --------------------------------#
@login_required
def assign_approver(request, model_type, pk, level):
    """Assign approver for a StarterSheet or CheckSheet"""
    if request.method == "POST":
        approver_id = request.POST.get("approver_id")

        if model_type == "startersheet":
            model = get_object_or_404(StarterSheet, pk=pk)
            redirect_url = "all_startersheet"
        elif model_type == "checksheet":
            model = get_object_or_404(CheckSheet, pk=pk)
            redirect_url = "all_checksheets"
        else:
            messages.error(request, "Invalid model type")
            return redirect("dashboard")  # Fallback to dashboard or another safe URL

        if level == "level1":
            approver = (
                get_object_or_404(CustomUser, pk=approver_id) if approver_id else None
            )
            model.level_1_approver = approver
            success_msg = f"Level 1 approver {approver.username if approver else 'none'} assigned to {model.name}"
        elif level == "level2":
            approver = (
                get_object_or_404(CustomUser, pk=approver_id) if approver_id else None
            )
            model.level_2_approver = approver
            success_msg = f"Level 2 approver {approver.username if approver else 'none'} assigned to {model.name}"

        model.save()
        messages.success(request, success_msg)

        return redirect(redirect_url)

    # Fallback redirect for GET requests
    if model_type == "startersheet":
        return redirect("all_startersheet")
    else:
        return redirect("all_checksheets")


# ------------------------------ for admin approvel yes or no  --------------------------------#


@login_required
def toggle_level3_approval(request, pk, model_type=None):
    """Toggle whether Level 3 (admin) approval is required for a StarterSheet or CheckSheet"""
    if request.method == "POST":
        # Determine model type from URL parameter if provided
        if model_type is None:
            # Try to determine from referer or other context
            referer = request.META.get("HTTP_REFERER", "")
            if "checksheet" in referer:
                model_type = "checksheet"
            else:
                model_type = "startersheet"  # Default to startersheet

        if model_type == "startersheet":
            model = get_object_or_404(StarterSheet, pk=pk)
            redirect_url = "all_startersheet"
        elif model_type == "checksheet":
            model = get_object_or_404(CheckSheet, pk=pk)
            redirect_url = "all_checksheets"
        else:
            messages.error(request, "Invalid model type")
            return redirect("dashboard")  # Fallback

        require_level_3 = request.POST.get("require_level_3") == "True"

        if model_type == "startersheet":
            model.require_level_3_approval = require_level_3
        else:  # checksheet
            model.require_level_3_approval = require_level_3

        model.save()

        status = "required" if require_level_3 else "not required"
        messages.success(request, f"Level 3 approval is now {status} for {model.name}")

        return redirect(redirect_url)

    # Default redirect for GET requests or if something went wrong
    return redirect(redirect_url)


# ------------------------------approvel hierarchy of checksheet  --------------------------------#


def get_checksheet_approval_hierarchy(request):
    # Extract request parameters
    checksheet_name = request.GET.get("name")
    username = request.GET.get("username")
    shift = request.GET.get("shift")
    date = request.GET.get("date")
    line = request.GET.get("line")  # Added line parameter

    # Validate required parameters
    if not all(
        [checksheet_name, username, shift, date, line]
    ):  # Added line to required params
        return JsonResponse({"error": "Missing required parameters"}, status=400)

    try:
        # First, get the user object by username
        try:
            user_obj = User.objects.get(username=username)
        except User.DoesNotExist:
            return JsonResponse({"error": f"User '{username}' not found"}, status=404)

        # Next, get the checksheet by name AND line
        try:
            checksheet_obj = CheckSheet.objects.get(name=checksheet_name, line=line)
        except CheckSheet.DoesNotExist:
            return JsonResponse(
                {
                    "error": f"Checksheet '{checksheet_name}' with line '{line}' not found"
                },
                status=404,
            )

        # Now fetch the filled checksheet with correct objects
        # Assuming timestamp field might be used for date filtering
        entries = FilledCheckSheet.objects.filter(
            checksheet=checksheet_obj,
            user=user_obj,
            shift=shift,
            line=line,  # Added line filter
            timestamp__date=datetime.strptime(
                date, "%Y-%m-%d"
            ).date(),  # Adjust format as needed
        )
        # Check if any entries were found
        if not entries.exists():
            return JsonResponse(
                {"error": "No matching checksheet entries found"}, status=404
            )

        # Consolidate the data as requested
        consolidated_data = {}
        for entry in entries:
            # Use only checksheet name, shift, and specific zones as key
            # since we're already filtering by user and date
            key = (entry.checksheet.name, entry.shift)
            if key not in consolidated_data:
                consolidated_data[key] = defaultdict(int)

            # Process status data
            if entry.status_data:
                for k, v in entry.status_data.items():
                    if v == "Yes":
                        consolidated_data[key][k] += 1
                    elif isinstance(v, int):
                        consolidated_data[key][k] += v
                    elif k == "completely_reject":
                        if isinstance(v, int):
                            consolidated_data[key][k] += v
                        else:
                            consolidated_data[key]["completely_reject"] += 1

        # Get the first entry for approval hierarchy information
        # (assuming all entries in the batch have the same approval flow)
        entry = entries.first()

        # Initialize the approval hierarchy
        approval_hierarchy = {}
        rejection_info = None

        # Level 1 approver info
        if entry.assigned_level_1_approver_id:
            try:
                level1_approver = User.objects.get(
                    id=entry.assigned_level_1_approver_id
                )
                approval_hierarchy["level1"] = {
                    "assigned_to": level1_approver.username,
                    "status": "Pending",
                    "timestamp": None,
                }

                # Check if approved by level 1
                if entry.level_1_approval_id:
                    level1_approver_action = User.objects.get(
                        id=entry.level_1_approval_id
                    )
                    approval_hierarchy["level1"].update(
                        {
                            "status": "Approved",
                            "action_by": level1_approver_action.username,
                            "timestamp": entry.level_1_approval_timestamp,
                        }
                    )
            except User.DoesNotExist:
                approval_hierarchy["level1"] = {
                    "assigned_to": "Unknown User",
                    "status": "Unknown",
                }

        # Level 2 approver info
        if entry.assigned_level_2_approver_id:
            try:
                level2_approver = User.objects.get(
                    id=entry.assigned_level_2_approver_id
                )
                approval_hierarchy["level2"] = {
                    "assigned_to": level2_approver.username,
                    "status": "Pending",
                    "timestamp": None,
                }

                # Check if approved by level 2
                if entry.level_2_approval_id:
                    level2_approver_action = User.objects.get(
                        id=entry.level_2_approval_id
                    )
                    approval_hierarchy["level2"].update(
                        {
                            "status": "Approved",
                            "action_by": level2_approver_action.username,
                            "timestamp": entry.level_2_approval_timestamp,
                        }
                    )
            except User.DoesNotExist:
                approval_hierarchy["level2"] = {
                    "assigned_to": "Unknown User",
                    "status": "Unknown",
                }

        # Level 3 approver info (if required)
        if entry.requires_level_3_approval:
            approval_hierarchy["level3"] = {
                "assigned_to": "Admin",
                "status": "Pending",
                "timestamp": None,
            }

            # Check if approved by level 3
            if entry.level_3_approval_id:
                try:
                    level3_approver_action = User.objects.get(
                        id=entry.level_3_approval_id
                    )
                    approval_hierarchy["level3"].update(
                        {
                            "status": "Approved",
                            "action_by": level3_approver_action.username,
                            "timestamp": entry.level_3_approval_timestamp,
                        }
                    )
                except User.DoesNotExist:
                    approval_hierarchy["level3"].update(
                        {
                            "action_by": "Unknown User",
                            "timestamp": entry.level_3_approval_timestamp,
                        }
                    )

        # Rejection info
        if entry.rejected_by_id:
            try:
                rejector = User.objects.get(id=entry.rejected_by_id)
                rejection_info = {
                    "rejected_by": rejector.username,
                    "reason": entry.rejection_reason,
                    "timestamp": entry.rejection_timestamp,
                }

                # Update the status of the relevant level to "Rejected"
                if entry.level_1_approval_id and not entry.level_2_approval_id:
                    # Rejected by Level 2
                    if "level2" in approval_hierarchy:
                        approval_hierarchy["level2"].update(
                            {
                                "status": "Rejected",
                                "action_by": rejector.username,
                                "timestamp": entry.rejection_timestamp,
                            }
                        )
                elif (
                    entry.level_2_approval_id
                    and not entry.level_3_approval_id
                    and entry.requires_level_3_approval
                ):
                    # Rejected by Level 3
                    if "level3" in approval_hierarchy:
                        approval_hierarchy["level3"].update(
                            {
                                "status": "Rejected",
                                "action_by": rejector.username,
                                "timestamp": entry.rejection_timestamp,
                            }
                        )
                else:
                    # Rejected by Level 1
                    if "level1" in approval_hierarchy:
                        approval_hierarchy["level1"].update(
                            {
                                "status": "Rejected",
                                "action_by": rejector.username,
                                "timestamp": entry.rejection_timestamp,
                            }
                        )
            except User.DoesNotExist:
                rejection_info = {
                    "rejected_by": "Unknown User",
                    "reason": entry.rejection_reason,
                    "timestamp": entry.rejection_timestamp,
                }

        # Convert datetime objects to ISO format strings for JSON serialization
        for level_key, level_data in approval_hierarchy.items():
            if level_data.get("timestamp"):
                level_data["timestamp"] = level_data["timestamp"].isoformat()

        if rejection_info and rejection_info.get("timestamp"):
            rejection_info["timestamp"] = rejection_info["timestamp"].isoformat()

        # Convert consolidated data to a format suitable for JSON
        consolidated_results = []
        for (checksheet_name, shift), counts in consolidated_data.items():
            consolidated_results.append(
                {
                    "checksheet_name": checksheet_name,
                    "shift": shift,
                    "counts": dict(counts),
                }
            )

        # Return the data as JSON
        return JsonResponse(
            {
                "approval_hierarchy": approval_hierarchy,
                "rejection_info": rejection_info,
                "consolidated_data": consolidated_results,
            }
        )

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# ------------------------------edit the checksheet count by admin in settings  --------------------------------#
@login_required
@csrf_exempt
def update_checksheet_errors(request):
    """
    Update the checksheet error counts in the database while preserving original timestamp
    and approval status.
    """
    try:
        data = json.loads(request.body)

        # Extract data from request
        line = data.get("line")
        checksheet_name = data.get("checksheet_name")
        shift = data.get("shift")
        date_str = data.get("date")
        error_values = data.get("error_values")

        # Validate required fields
        if not all([line, checksheet_name, shift, date_str, error_values]):
            return JsonResponse({"error": "Missing required fields"}, status=400)

        # Parse date
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return JsonResponse(
                {"error": "Invalid date format. Use YYYY-MM-DD"}, status=400
            )

        # Find the checksheet by name and line
        try:
            checksheet = CheckSheet.objects.get(name=checksheet_name, line=line)
        except CheckSheet.DoesNotExist:
            return JsonResponse(
                {
                    "error": f"Checksheet '{checksheet_name}' for line '{line}' not found"
                },
                status=404,
            )

        # Use transaction to ensure data integrity
        with transaction.atomic():
            # Find all entries for this checksheet, shift, and date
            entries = FilledCheckSheet.objects.filter(
                checksheet_id=checksheet.id, shift=shift, timestamp__date=date_obj
            )

            if not entries.exists():
                return JsonResponse(
                    {"error": "No entries found for the specified criteria"}, status=404
                )

            # Get the first entry to preserve its attributes
            first_entry = entries.first()

            # Create a new consolidated entry with ALL attributes preserved from first_entry
            new_entry = FilledCheckSheet(
                status_data=error_values,
                line=first_entry.line,
                timestamp=first_entry.timestamp,
                shift=shift,
                checksheet_id=checksheet.id,
                user_id=first_entry.user_id,
                send_acknowledgment=first_entry.send_acknowledgment,
                approval_status=first_entry.approval_status,
                assigned_level_1_approver_id=first_entry.assigned_level_1_approver_id,
                assigned_level_2_approver_id=first_entry.assigned_level_2_approver_id,
                level_1_approval_id=first_entry.level_1_approval_id,
                level_1_approval_timestamp=first_entry.level_1_approval_timestamp,
                level_2_approval_id=first_entry.level_2_approval_id,
                level_2_approval_timestamp=first_entry.level_2_approval_timestamp,
                level_3_approval_id=first_entry.level_3_approval_id,
                level_3_approval_timestamp=first_entry.level_3_approval_timestamp,
                rejected_by_id=first_entry.rejected_by_id,
                rejection_reason=first_entry.rejection_reason,
                rejection_timestamp=first_entry.rejection_timestamp,
                requires_level_3_approval=first_entry.requires_level_3_approval,
            )

            # Delete all existing entries only after successfully creating the new one
            entries.delete()

            # Save the new entry
            new_entry.save()

        return JsonResponse(
            {
                "message": "Checksheet error counts updated successfully",
                "data": {
                    "line": line,
                    "checksheet_name": checksheet_name,
                    "shift": shift,
                    "date": date_str,
                    "timestamp": first_entry.timestamp.strftime("%Y-%m-%d %H:%M:%S.%f"),
                    "status_data": error_values,
                    "approval_status": first_entry.approval_status,
                },
            },
            status=200,
        )

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def get_checksheet_data(request):
    """
    Get the checksheet data for editing, including approval status information.
    """
    try:
        # Extract query parameters
        line = request.GET.get("line")
        checksheet_name = request.GET.get("checksheet_name")
        shift = request.GET.get("shift")
        date_str = request.GET.get("date")

        # Validate required fields
        if not all([line, checksheet_name, shift, date_str]):
            return JsonResponse({"error": "Missing required parameters"}, status=400)

        # Parse date
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return JsonResponse(
                {"error": "Invalid date format. Use YYYY-MM-DD"}, status=400
            )

        # Find the checksheet by name and line
        try:
            checksheet = CheckSheet.objects.get(name=checksheet_name, line=line)
        except CheckSheet.DoesNotExist:
            return JsonResponse(
                {
                    "error": f"Checksheet '{checksheet_name}' for line '{line}' not found"
                },
                status=404,
            )

        # Find entries for this checksheet, shift, and date
        entries = FilledCheckSheet.objects.filter(
            checksheet_id=checksheet.id, shift=shift, timestamp__date=date_obj
        )

        if not entries.exists():
            return JsonResponse(
                {"error": "No data found for the specified criteria"}, status=404
            )

        # Get the first entry to fetch approval information
        first_entry = entries.first()

        # Consolidate the status data (based on your data consolidation logic)
        consolidated_data = {}
        for entry in entries:
            for k, v in entry.status_data.items():
                if v == "Yes":
                    consolidated_data[k] = consolidated_data.get(k, 0) + 1
                elif isinstance(v, int):
                    consolidated_data[k] = consolidated_data.get(k, 0) + v

        # Collect approval information
        approval_info = {
            "approval_status": first_entry.approval_status,
            "requires_level_3_approval": first_entry.requires_level_3_approval,
            "level_1_approved": bool(first_entry.level_1_approval_id),
            "level_2_approved": bool(first_entry.level_2_approval_id),
            "level_3_approved": bool(first_entry.level_3_approval_id),
            "rejected": bool(first_entry.rejected_by_id),
            "rejection_reason": first_entry.rejection_reason,
        }

        # Return the consolidated data with approval info
        return JsonResponse(
            {
                "line": line,
                "checksheet_name": checksheet_name,
                "shift": shift,
                "date": date_str,
                "timestamp": first_entry.timestamp.strftime("%Y-%m-%d %H:%M:%S.%f"),
                "status_data": consolidated_data,
                "username": request.user.username,
                "approval_info": approval_info,
            }
        )

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# Add a new view to get checksheets by line
@login_required
def get_checksheets_by_line(request):
    """
    Get all checksheets for a specific line
    """
    try:
        line = request.GET.get("line")

        if not line:
            return JsonResponse({"error": "Line parameter is required"}, status=400)

        checksheets = CheckSheet.objects.filter(line=line)

        # Convert to list of dict for JSON response
        checksheet_list = []
        for checksheet in checksheets:
            checksheet_list.append(
                {
                    "id": checksheet.id,
                    "name": checksheet.name,
                    "line": checksheet.line,
                    "requires_level_3_approval": checksheet.require_level_3_approval,
                }
            )

        return JsonResponse({"checksheets": checksheet_list})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# Add a new view to get all checksheets
@login_required
def get_all_checksheets(request):
    """
    Get all checksheets with their line information
    """
    try:
        checksheets = CheckSheet.objects.all()

        # Convert to list of dict for JSON response
        checksheet_list = []
        for checksheet in checksheets:
            checksheet_list.append(
                {
                    "id": checksheet.id,
                    "name": checksheet.name,
                    "line": checksheet.line,
                    "requires_level_3_approval": checksheet.require_level_3_approval,
                }
            )

        return JsonResponse({"checksheets": checksheet_list})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@login_required
def shiftpage(request):
    shift_instance = Shifttime.objects.first()

    if request.method == "POST":
        shift_A_start = request.POST.get("shift_A_start")
        shift_A_end = request.POST.get("shift_A_end")
        shift_B_start = request.POST.get("shift_B_start")
        shift_B_end = request.POST.get("shift_B_end")

        if shift_instance:
            # Update existing record
            shift_instance.shift_A_start = shift_A_start
            shift_instance.shift_A_end = shift_A_end
            shift_instance.shift_B_start = shift_B_start
            shift_instance.shift_B_end = shift_B_end
            shift_instance.save()
        else:
            # If no record exists, create one
            Shifttime.objects.create(
                shift_A_start=shift_A_start,
                shift_A_end=shift_A_end,
                shift_B_start=shift_B_start,
                shift_B_end=shift_B_end,
            )
        messages.success(request, "Shift Saved", extra_tags="shift")
        return redirect("setting_view")  # Redirect to refresh the page

    return render(request, "checksheet/shift_page.html", {"shift": shift_instance})


@require_http_methods(["GET"])
def get_shift_data(request):
    """
    Initial endpoint to get all shift times at once
    """
    IST = pytz.timezone("Asia/Kolkata")
    current_time = now().astimezone(IST).time()

    try:
        shift_times = Shifttime.objects.first()

        # Determine which shift should be active based on the current time
        current_shift = None
        if shift_times.shift_A_start <= current_time < shift_times.shift_A_end:
            current_shift = "A"
        elif shift_times.shift_B_start <= current_time < shift_times.shift_B_end:
            current_shift = "B"

        # Save current shift in session
        request.session["current_shift"] = current_shift

        return JsonResponse(
            {
                "shift_times": {
                    "shift_a_start": shift_times.shift_A_start.strftime("%H:%M:%S"),
                    "shift_a_end": shift_times.shift_A_end.strftime("%H:%M:%S"),
                    "shift_b_start": shift_times.shift_B_start.strftime("%H:%M:%S"),
                    "shift_b_end": shift_times.shift_B_end.strftime("%H:%M:%S"),
                },
                "current_shift": current_shift,
                "current_time": current_time.strftime("%H:%M:%S"),
            }
        )

    except Shifttime.DoesNotExist:
        return JsonResponse(
            {
                "error": "Shift times not configured",
                "current_time": current_time.strftime("%H:%M:%S"),
            },
            status=404,
        )


# ----------------- shift function to change shift ------------#
@require_http_methods(["POST"])
def verify_shift_change(request):
    """
    Endpoint to verify if a shift change should result in a logout
    """
    IST = pytz.timezone("Asia/Kolkata")
    current_time = now().astimezone(IST).time()

    try:
        # Parse the request body
        data = json.loads(request.body)
        client_shift = data.get("client_shift")

        # Get user's current stored shift
        session_shift = request.session.get("current_shift")

        shift_times = Shifttime.objects.first()

        # Determine which shift should be active based on the current time
        server_shift = None
        if shift_times.shift_A_start <= current_time < shift_times.shift_A_end:
            server_shift = "A"
        elif shift_times.shift_B_start <= current_time < shift_times.shift_B_end:
            server_shift = "B"

        # Update the session with the current shift
        request.session["current_shift"] = server_shift

        # Determine if logout is needed
        should_logout = False

        # Case 1: Client thinks it's end of shift B
        if client_shift == "B" and server_shift is None:
            # Verify we're close to shift B end
            if is_time_near(current_time, shift_times.shift_B_end, minutes=3):
                should_logout = True

        # Case 2: Client thinks it's end of shift A but server says B
        elif session_shift == "A" and server_shift == "B":
            # Verify we're close to shift A end
            if is_time_near(current_time, shift_times.shift_A_end, minutes=3):
                should_logout = True

        # Case 3: Shift mismatch between session and server (rare case)
        elif session_shift and server_shift and session_shift != server_shift:
            should_logout = True

        return JsonResponse(
            {
                "should_logout": should_logout,
                "current_shift": server_shift,
                "server_time": current_time.strftime("%H:%M:%S"),
            }
        )

    except (Shifttime.DoesNotExist, json.JSONDecodeError) as e:
        return JsonResponse(
            {
                "error": str(e),
                "should_logout": False,
                "current_time": current_time.strftime("%H:%M:%S"),
            },
            status=400,
        )


# ------------to show countdown of shift ending -------------#


def is_time_near(current_time, target_time, minutes=3):
    """
    Helper function to check if current time is within X minutes of target time
    """
    if not current_time or not target_time:
        return False

    # Calculate the time range
    import datetime

    # Convert to datetime.datetime for easier manipulation
    base_date = datetime.datetime.today().date()
    current_datetime = datetime.datetime.combine(base_date, current_time)
    target_datetime = datetime.datetime.combine(base_date, target_time)

    # Calculate time difference in minutes
    time_diff = abs((current_datetime - target_datetime).total_seconds()) / 60

    return time_diff <= minutes


# ------------------------setting page -----------------#
@login_required
def setting_view(request):
    """
    Combined view for the tabbed interface that includes shift page, error editor, rejection alert, target, and reject reason
    """
    # Get shift data
    shift_instance = Shifttime.objects.first()

    # Get checksheets for the error editor
    checksheets = CheckSheet.objects.all()

    # Get RejectionAlertConfig, handle case where it doesn't exist
    config = RejectionAlertConfig.objects.first()
    # Get recipients from config, default to empty list if config doesn't exist
    recipients = config.get_alert_recipients() if config else []
    print("Recipients:", recipients)

    # Get ProductionTarget, handle case where it doesn't exist
    production_target = ProductionTarget.objects.first() or ProductionTarget(target_value=0)

    # Get Reject Reasons
    reject_reasons = RejectReason.objects.all()

    context = {
        "checksheets": checksheets,
        "config": config,
        "shift": shift_instance,
        "recipients": recipients,  # Add recipients to context
        "recipient_count": len(recipients) or 1,  # Add recipient_count for consistency
        "production_target": production_target,
        "reject_reasons": reject_reasons,
        "users": CustomUser.objects.all(),
        "reason_count": 0,
    }

    return render(request, "checksheet/settings.html", context)

@login_required
def save_reject_reasons(request):
    """
    API endpoint to save (create or update) reject reasons
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            reasons = data.get('reasons', [])
            response_reasons = []

            for reason_data in reasons:
                reason_id = reason_data.get('id')
                reason_text = reason_data.get('reason')

                if not reason_text:
                    continue

                if reason_id:  # Update existing reason
                    try:
                        reason_obj = RejectReason.objects.get(id=reason_id)
                        reason_obj.reason = reason_text
                        reason_obj.save()
                        response_reasons.append({'id': reason_obj.id, 'reason': reason_obj.reason})
                    except RejectReason.DoesNotExist:
                        continue
                else:  # Create new reason
                    reason_obj = RejectReason.objects.create(reason=reason_text)
                    response_reasons.append({'id': reason_obj.id, 'reason': reason_obj.reason})

            return JsonResponse({'reasons': response_reasons}, status=200)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
def reject_reason_config(request):
    """
    Handle the reject reason configuration form submission
    """
    if request.method == 'POST':
        try:
            # Handle new reasons
            new_reasons = request.POST.getlist('reject_reason')
            
            # Handle existing reasons - Updated to match new form structure
            existing_reasons = []
            for key, value in request.POST.items():
                if key.startswith('edit_reason_') and value.strip():
                    reason_id = key.replace('edit_reason_', '')
                    existing_reasons.append((reason_id, value.strip()))
            
            print(f"New reasons: {new_reasons}")
            print(f"Existing reasons: {existing_reasons}")
            
            # Update existing reasons
            for reason_id, new_reason_text in existing_reasons:
                try:
                    reason = RejectReason.objects.get(id=reason_id)
                    reason.reason = new_reason_text
                    reason.save()
                    print(f"Updated reason {reason_id}: {new_reason_text}")
                except RejectReason.DoesNotExist:
                    messages.error(request, f"Reason with ID {reason_id} not found.")
                    continue
                except Exception as e:
                    messages.error(request, f"Error updating reason {reason_id}: {str(e)}")
                    continue

            # Add new reasons only if they are provided
            new_reasons_created = 0
            for reason in new_reasons:
                if reason.strip():
                    try:
                        RejectReason.objects.create(reason=reason.strip())
                        new_reasons_created += 1
                        print(f"Created new reason: {reason.strip()}")
                    except Exception as e:
                        messages.error(request, f"Error creating reason '{reason}': {str(e)}")
                        continue

            # Success message
            success_parts = []
            if existing_reasons:
                success_parts.append(f"{len(existing_reasons)} existing reason(s) updated")
            if new_reasons_created > 0:
                success_parts.append(f"{new_reasons_created} new reason(s) created")
            
            if success_parts:
                messages.success(request, f"Reject reasons updated successfully! {', '.join(success_parts)}")
            else:
                messages.info(request, "No changes were made.")
                
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            print(f"Error in reject_reason_config: {str(e)}")

        return redirect('setting_view')

    return redirect('setting_view')

@login_required
def delete_reject_reason(request, reason_id):
    """
    API endpoint to delete a reject reason
    """
    from django.http import JsonResponse
    try:
        reason = RejectReason.objects.get(id=reason_id)
        reason.delete()
        return JsonResponse({"status": "success", "message": "Reject reason deleted successfully"})
    except RejectReason.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Reject reason not found"}, status=404)
    except Exception as e:
        return JsonResponse({"status": "error", "message": str(e)}, status=500)

@login_required
def production_target_view(request):
    """
    View to handle production target form submission
    """
    if request.method == "POST":
        target_value = request.POST.get("production_target")
        try:
            target_value = int(target_value)
            if target_value < 0:
                messages.error(request, "Production target must be a positive number.")
            else:
                # Update or create the production target
                production_target, created = ProductionTarget.objects.get_or_create(
                    id=1, defaults={"target_value": target_value}
                )
                if not created:
                    production_target.target_value = target_value
                    production_target.save()
                messages.success(request, "Production target saved successfully!")
        except ValueError:
            messages.error(request, "Please enter a valid number for the production target.")
        
        return redirect(setting_view)
    
    return redirect(setting_view)



# --------------rejection alert functions------------------#
def rejection_alert_config(request):
    config, created = RejectionAlertConfig.objects.get_or_create(pk=1)

    if request.method == "POST":
        try:
            rejection_threshold = int(request.POST.get("rejection_threshold", 2))
            recipient_count = int(request.POST.get("recipient_count", 1))
            recipients = []

            for i in range(recipient_count):
                user_id = request.POST.get(f"user_id_{i}")
                phone_number = request.POST.get(f"phone_number_{i}")
                percentage = request.POST.get(f"percentage_{i}")
                if user_id and phone_number and percentage:
                    # Preserve existing last_sms_sent if recipient exists, else set to None
                    existing_recipients = config.get_alert_recipients()
                    last_sms_sent = None
                    for existing in existing_recipients:
                        if existing.get("user_id") == user_id:
                            last_sms_sent = existing.get("last_sms_sent")
                            break
                    recipients.append({
                        "user_id": user_id,
                        "phone_number": phone_number,
                        "percentage": float(percentage),
                        "last_sms_sent": last_sms_sent  # Preserve or initialize
                    })

            # Validate individual percentages
            for recipient in recipients:
                percentage = float(recipient["percentage"])
                if percentage <= 0 or percentage > 100:
                    messages.error(request, f"Percentage for user {recipient['user_id']} must be between 0 and 100.")
                    return redirect("rejection_alert_config")

            config.rejection_threshold = rejection_threshold
            config.set_alert_recipients(recipients)
            config.save()

            messages.success(request, "Alert configuration saved successfully!")
        except Exception as e:
            print(f"Error saving configuration: {str(e)}")
            messages.error(request, f"Error saving configuration: {str(e)}")

        return redirect("rejection_alert_config")

    context = {
        "config": config,
        "users": CustomUser.objects.all(),
        "recipients": config.get_alert_recipients(),
        "recipient_count": len(config.get_alert_recipients()) or 1,
    }
    return render(request, "checksheet/settings.html", context)

@login_required
def delete_recipient(request, index):
    """
    API endpoint to delete a specific recipient from the RejectionAlertConfig by index.
    """
    if request.method != "DELETE":
        return JsonResponse({"error": "Method not allowed"}, status=405)

    try:
        config = RejectionAlertConfig.objects.first()
        if not config:
            return JsonResponse({"error": "No configuration found"}, status=404)

        recipients = config.get_alert_recipients()
        if index < 0 or index >= len(recipients):
            return JsonResponse({"error": "Invalid recipient index"}, status=400)

        # Remove the recipient at the specified index
        recipients.pop(index)
        config.set_alert_recipients(recipients)
        config.save()

        return JsonResponse({"message": "Recipient deleted successfully"}, status=200)
    except Exception as e:
        return JsonResponse({"error": f"Error deleting recipient: {str(e)}"}, status=500)